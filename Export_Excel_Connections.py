import os
import sys
import argparse
import zipfile
import time
from typing import List, Tuple

# Try importing openpyxl; inform user clearly if missing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
except ImportError as e:
    print("Missing dependency: openpyxl. Please install it (pip install openpyxl).")
    sys.exit(1)


DEFAULT_ROOT = r"C:\\Users\\giuseppe.tanda\\Desktop\\doValue\\Report excel"
DEFAULT_OUTPUT = "Report/Excel_Connections_Report.xlsx"
EXCEL_CONN_PATH = "xl/connections.xml"


def find_connections_in_xlsx(root_dir: str, verbose: bool = True) -> List[Tuple[str, str]]:
    """Traverse root_dir, find .xlsx files and extract xl/connections.xml content.

    Returns a list of tuples: (file_name, connections_xml_content)
    Only includes files that contain the connections.xml entry.
    """
    results: List[Tuple[str, str]] = []
    processed_xlsx = 0
    found_count = 0

    # First pass: count total .xlsx to enable 1/N progress display
    total_xlsx = 0
    for dirpath, dirnames, filenames in os.walk(root_dir):
        for fname in filenames:
            if fname.lower().endswith(".xlsx"):
                total_xlsx += 1
    if verbose:
        print(f"Total .xlsx files to process: {total_xlsx}")
    for dirpath, dirnames, filenames in os.walk(root_dir):
        if verbose:
            print(f"Scanning folder: {dirpath}")
        for fname in filenames:
            if not fname.lower().endswith(".xlsx"):
                continue
            full_path = os.path.join(dirpath, fname)
            processed_xlsx += 1
            if verbose:
                print(f"[{processed_xlsx}/{total_xlsx}] Processing: {fname}")
            try:
                with zipfile.ZipFile(full_path, "r") as zf:
                    if EXCEL_CONN_PATH in zf.namelist():
                        data = zf.read(EXCEL_CONN_PATH)
                        # Decode safely; Excel files are typically UTF-8 here
                        xml_text = data.decode("utf-8", errors="replace")
                        # Keep only file name as requested
                        results.append((fname, xml_text))
                        found_count += 1
            except zipfile.BadZipFile:
                # Skip corrupted or non-standard xlsx files
                if verbose:
                    print(f"WARN: Bad or corrupted xlsx skipped: {full_path}")
                continue
            except Exception as ex:
                # Be resilient; skip unexpected errors on individual files
                if verbose:
                    print(f"WARN: Unexpected error reading {full_path}: {ex}")
                continue
    if verbose:
        print(f"Scan complete: processed {processed_xlsx}/{total_xlsx} xlsx files, found {found_count} with connections.xml.")
    return results


def write_report(rows: List[Tuple[str, str]], output_path: str) -> str:
    """Write results to an Excel file with two columns: File, Connections XML.

    Caps cell content to Excel's limit (32767 chars) to avoid write errors.
    Returns the absolute output path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Connections"

    ws.cell(row=1, column=1, value="File")
    ws.cell(row=1, column=2, value="Connections XML")

    wrap = Alignment(wrap_text=True, vertical="top")
    max_cell_len = 32767

    for idx, (fname, xml_text) in enumerate(rows, start=2):
        # Enforce Excel cell text length limit
        cell_text = xml_text if len(xml_text) <= max_cell_len else (xml_text[:max_cell_len - 15] + "... [TRUNCATED]")
        ws.cell(row=idx, column=1, value=fname)
        c2 = ws.cell(row=idx, column=2, value=cell_text)
        c2.alignment = wrap

    # Basic column width tuning
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 120

    # Ensure parent directory exists
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
    return os.path.abspath(output_path)


def main():
    parser = argparse.ArgumentParser(description="Scan xlsx files for xl/connections.xml and export report.")
    parser.add_argument("root", nargs="?", default=DEFAULT_ROOT, help="Root folder to scan (default: configured path)")
    parser.add_argument("--out", dest="out", default=DEFAULT_OUTPUT, help="Output Excel file path (default: Report/Excel_Connections_Report.xlsx)")
    parser.add_argument("--quiet", action="store_true", help="Reduce logging output")
    args = parser.parse_args()

    root_dir = args.root
    output_path = args.out

    if not os.path.exists(root_dir):
        print(f"Root path does not exist: {root_dir}")
        sys.exit(2)

    print(f"Scanning for .xlsx under: {root_dir}")
    start = time.time()
    rows = find_connections_in_xlsx(root_dir, verbose=not args.quiet)
    elapsed = time.time() - start
    print(f"Found {len(rows)} files with xl/connections.xml (elapsed {elapsed:.1f}s)")

    if not rows:
        print("No connections.xml found in any xlsx.")
        # Still create an empty report with headers
        abs_out = write_report([], output_path)
        print(f"Report written: {abs_out}")
        return

    abs_out = write_report(rows, output_path)
    print(f"Report written: {abs_out}")


if __name__ == "__main__":
    main()
