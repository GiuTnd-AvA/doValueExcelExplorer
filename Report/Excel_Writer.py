import os
import re
from typing import Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook

class ExcelWriter:
    
    def __init__(self, folder_path, file_name):
        self.folder_path = folder_path
        self.file_name = file_name
        # Tracks whether the file has been initialized in the current run
        self._initialized = False
        # Resolved output path to keep sheets in the same file across writes
        self._resolved_output_path = None

    def write_excel(self, columns, data, sheet_name='Sheet1'):
        # Ensure output directory exists
        if self.folder_path and not os.path.isdir(self.folder_path):
            os.makedirs(self.folder_path, exist_ok=True)

        base_output_path = os.path.join(self.folder_path, self.file_name)
        if self._resolved_output_path is None:
            self._resolved_output_path = base_output_path
        output_path = self._resolved_output_path
        df = pd.DataFrame(data, columns=columns)

        # Sanitize DataFrame to remove characters illegal for openpyxl
        # Remove ASCII control chars except tab(\x09), newline(\x0A), carriage return(\x0D)
        illegal_chars = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

        def _clean(val):
            if isinstance(val, str):
                return illegal_chars.sub("", val)
            return val

        # Clean each cell value (applymap is supported for DataFrame element-wise ops)
        df = df.applymap(_clean)

        # Overwrite file on first write of this instance; append thereafter
        if not self._initialized:
            # First write in this run: start fresh (overwrite file)
            mode = 'w'
            # If a previous file exists, starting with 'w' will overwrite it
            self._initialized = True
        else:
            # Subsequent writes in the same run: append sheets
            mode = 'a'

        writer_kwargs = dict(engine='openpyxl', mode=mode)
        if mode == 'a':
            writer_kwargs['if_sheet_exists'] = 'replace'

        # Also ensure sheet_name is clean and within Excel limits
        clean_sheet_name = illegal_chars.sub("", sheet_name)[:31] or "Sheet1"

        try:
            with pd.ExcelWriter(output_path, **writer_kwargs) as writer:
                df.to_excel(writer, index=False, sheet_name=clean_sheet_name)
        except PermissionError:
            # If the target file is locked (e.g., opened in Excel), fall back to a timestamped filename
            import datetime
            ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            alt_name = os.path.splitext(self.file_name)[0] + f"_{ts}" + os.path.splitext(self.file_name)[1]
            self._resolved_output_path = os.path.join(self.folder_path, alt_name)
            output_path = self._resolved_output_path
            # On first write, we still want 'w'; for subsequent writes, keep 'a'
            # Always start a new file in write mode for the fallback path
            writer_kwargs = dict(engine='openpyxl', mode='w')
            with pd.ExcelWriter(output_path, **writer_kwargs) as writer:
                df.to_excel(writer, index=False, sheet_name=clean_sheet_name)
            # Mark as initialized so subsequent writes append to the same file
            self._initialized = True


# -----------------------------------------------------------------------------
# Utility helpers for writing large outputs split across multiple Excel files.
# -----------------------------------------------------------------------------
EXCEL_MAX_ROWS: int = 1_048_576  # Excel per-sheet limit, including header row
_DATA_ROWS_PER_SHEET: int = EXCEL_MAX_ROWS - 1  # accounting for header row


def _derive_part_path(base_path: str, part_index: int) -> str:
    """Return a file path with a _partN suffix before extension.

    For part_index == 1, returns base_path to keep the first file name intact.
    """
    root, ext = os.path.splitext(base_path)
    if not ext:
        ext = ".xlsx"
    if part_index == 1:
        return root + ext
    return f"{root}_part{part_index}{ext}"


def write_dataframe_split_across_files(
    df: pd.DataFrame,
    base_output_path: str,
    sheet_name: str = "Sheet1",
) -> List[str]:
    """Write a DataFrame to one or more Excel files, splitting by row limit.

    Returns the list of written file paths.
    """
    if df is None:
        return []

    total_rows = len(df)
    written: List[str] = []
    if total_rows <= _DATA_ROWS_PER_SHEET:
        out_path = _derive_part_path(base_output_path, 1)
        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        written.append(out_path)
        return written

    part = 1
    for start in range(0, total_rows, _DATA_ROWS_PER_SHEET):
        end = min(start + _DATA_ROWS_PER_SHEET, total_rows)
        chunk = df.iloc[start:end]
        out_path = _derive_part_path(base_output_path, part)
        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            chunk.to_excel(writer, index=False, sheet_name=sheet_name)
        written.append(out_path)
        part += 1
    return written


def write_rows_split_across_files(
    headers: Sequence[str],
    rows: Iterable[Sequence],
    base_output_path: str,
    sheet_name: str = "Sheet1",
    column_widths: Optional[Sequence[int]] = None,
) -> List[str]:
    """Write rows (list of sequences) to one or more Excel files with openpyxl.

    - headers: sequence of header strings
    - rows: iterable of row sequences matching headers length
    - base_output_path: path template; _partN suffix will be added as needed
    - sheet_name: name of the worksheet
    - column_widths: optional widths to set per column (1-based)

    Returns the list of written file paths.
    """
    headers = list(headers)
    # Materialize rows if it's a generator to compute length safely
    rows_list: List[Sequence] = list(rows)
    total_rows = len(rows_list)

    def _write_chunk(chunk_rows: List[Sequence], out_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31] or "Sheet1"

        # Write header (bold)
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # Write data rows
        for r in chunk_rows:
            ws.append(list(r))

        # Column widths if provided
        if column_widths:
            from openpyxl.utils import get_column_letter

            for i, w in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        wb.save(out_path)

    written: List[str] = []
    if total_rows <= _DATA_ROWS_PER_SHEET:
        out_path = _derive_part_path(base_output_path, 1)
        _write_chunk(rows_list, out_path)
        written.append(out_path)
        return written

    part = 1
    start = 0
    while start < total_rows:
        end = min(start + _DATA_ROWS_PER_SHEET, total_rows)
        out_path = _derive_part_path(base_output_path, part)
        _write_chunk(rows_list[start:end], out_path)
        written.append(out_path)
        start = end
        part += 1
    return written