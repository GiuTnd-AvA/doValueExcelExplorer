import os
from typing import List, Optional

try:
    import pyodbc
except Exception:
    pyodbc = None  # type: ignore

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore

try:
    import pandas as pd
except Exception:
    pd = None  # type: ignore

# -----------------------------------------------------------------------------
# Configurazione: inserisci qui i percorsi e i parametri di connessione.
# -----------------------------------------------------------------------------
INPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\Selects.xlsx"
OUTPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\Esiti_Select.xlsx"
SQL_SERVER: str = "EPCP3"  # se istanza nominata: "EPCP3\\ISTANZA"
SQL_DATABASE: str = "master"
# Proveremo questi driver in ordine.
ODBC_DRIVERS: List[str] = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
]
TRUSTED_CONNECTION: bool = True
SQL_USERNAME: Optional[str] = None  # usato se TRUSTED_CONNECTION=False
SQL_PASSWORD: Optional[str] = None  # usato se TRUSTED_CONNECTION=False
# Timeout in secondi per ogni query
QUERY_TIMEOUT: int = 60
# Opzioni di cifratura/Trust (ODBC 18 abilita Encrypt by default). Regola se necessario.
ODBC_ENCRYPT_OPTS: str = "Encrypt=no;TrustServerCertificate=yes;"


class SelectsExecutor:
    """Legge SELECT da Excel (colonna A) ed esegue ciascuna sul server SQL indicato.

    Output: un Excel con due colonne: Select | Errore
    Se la SELECT esegue correttamente, 'Errore' rimane vuoto.
    """

    def __init__(self, input_excel: str, output_excel: str):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if pd is None:
            raise RuntimeError("pandas non installato. Installa 'pip install pandas openpyxl'.")
        if pyodbc is None:
            raise RuntimeError("pyodbc non installato. Installa 'pip install pyodbc'.")
        if not input_excel:
            raise ValueError("Percorso Excel di input non valorizzato.")
        self.input_excel = input_excel
        self.output_excel = output_excel or os.path.join(os.path.dirname(input_excel) or os.getcwd(), "Esiti_Select.xlsx")

    def _read_selects(self) -> List[str]:
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(min_row=1, max_col=1, values_only=True))
            selects: List[str] = []
            first_val = None
            if rows:
                first_cell = rows[0][0]
                first_val = str(first_cell).strip() if first_cell is not None else ""
            for r in rows:
                val = r[0]
                if val is None:
                    continue
                s = str(val).strip()
                if not s:
                    continue
                low = s.lower()
                # Accetta solo SELECT/CTE; ignora header tipo "conn"
                if low.startswith("select") or low.startswith("with"):
                    selects.append(s)
            # Se la prima riga è un'intestazione (ad es. "Select"), rimuovila
            if selects and first_val and first_val.lower() in ("select", "query", "sql"):
                selects = selects[1:]
            return selects
        finally:
            wb.close()

    def _build_conn_str(self) -> str:
        last_error: Optional[Exception] = None
        # Trova il primo driver disponibile tra quelli elencati
        for drv in ODBC_DRIVERS:
            try:
                conn_str = f"DRIVER={{{drv}}};SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
                if TRUSTED_CONNECTION:
                    conn_str += "Trusted_Connection=yes;"
                else:
                    if not SQL_USERNAME or not SQL_PASSWORD:
                        raise RuntimeError("Imposta SQL_USERNAME e SQL_PASSWORD oppure usa Trusted_Connection.")
                    conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                # Opzioni utili con ODBC 17/18
                conn_str += ODBC_ENCRYPT_OPTS
                # Proviamo una connessione veloce per validare il driver
                conn = pyodbc.connect(conn_str, timeout=3)
                conn.close()
                return conn_str
            except Exception as e:
                last_error = e
                continue
        # Se siamo qui, nessun driver ha funzionato
        raise RuntimeError(f"Nessun driver ODBC valido trovato. Ultimo errore: {last_error}")

    def _execute_select(self, conn, sql: str) -> Optional[str]:
        """Esegue una singola SELECT. Ritorna None se ok, altrimenti messaggio di errore."""
        try:
            cursor = conn.cursor()
            # rimosso: cursor.timeout (non esiste)
            cursor.execute(sql)
            # Non è richiesto leggere i risultati, vogliamo solo validare l'esecuzione
            # Recuperiamo un record al massimo per forzare eventuali errori di sintassi
            try:
                cursor.fetchmany(1)
            except Exception:
                # Se la query non produce result set (es. SET), ignoriamo la fetch
                pass
            return None
        except Exception as e:
            return str(e)

    def run(self) -> str:
        selects = self._read_selects()
        if not selects:
            raise RuntimeError("Nessuna SELECT trovata nell'Excel di input.")

        conn_str = self._build_conn_str()
        results: List[List[str]] = []
        conn = None
        try:
            conn = pyodbc.connect(conn_str, timeout=QUERY_TIMEOUT)
            for s in selects:
                err = self._execute_select(conn, s)
                results.append([s, "" if err is None else err])
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

        # Scrivi output
        out_dir = os.path.dirname(self.output_excel)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        df = pd.DataFrame(results, columns=["Select", "Errore"])
        with pd.ExcelWriter(self.output_excel, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name='Esiti')
        return self.output_excel


if __name__ == "__main__":
    if not INPUT_EXCEL_PATH:
        raise SystemExit("Imposta INPUT_EXCEL_PATH a inizio file.")
    executor = SelectsExecutor(INPUT_EXCEL_PATH, OUTPUT_EXCEL_PATH or "")
    out_path = executor.run()
    print(f"Esiti scritti in: {out_path}")
