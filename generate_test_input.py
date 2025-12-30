from openpyxl import Workbook
import os

BASE = os.path.abspath(os.path.dirname(__file__))
empty_path = os.path.join(BASE, 'empty.xlsx')
input_path = os.path.join(BASE, 'Connessioni assenti da verificare.xlsx')

# Create empty.xlsx
wb = Workbook()
wb.save(empty_path)

# Create input Excel listing the empty.xlsx path in Sheet1 column A
wb2 = Workbook()
ws = wb2.active
ws.title = 'Sheet1'
ws['A1'] = empty_path
wb2.save(input_path)
print('Created:', empty_path)
print('Created:', input_path)
