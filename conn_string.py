import os
import pandas as pd
from openpyxl import load_workbook
import zipfile
import xml.etree.ElementTree as ET
import re
 
def full_path(path, file_name):
    return file_name  # Funzione placeholder, non usata
folder_path = r"C:\Users\giuseppe.tanda\Desktop\doValue"

# Trova tutti i file .xlsx nella cartella radice e sottocartelle
files_name = []
for root, dirs, files in os.walk(folder_path):
    for f in files:
        if f.endswith(".xlsx") and not f.startswith("~$"):
            file_path = os.path.join(root, f)
            print(f"Trovato file: {file_path} (Cartella: {os.path.basename(root)})")
            files_name.append(file_path)


 
excel_file_list_df = []
 
for file in files_name:
    excel_df = pd.read_excel(file, engine="openpyxl")
    excel_df["file_origine"] = os.path.basename(file)
    folder_rel = os.path.relpath(os.path.dirname(file), folder_path)
    excel_df["folder_name"] = folder_rel
    excel_file_list_df.append(excel_df)
 

# # Estraggo connessioni esterne
# def extract_connections(zip_path):
#     connections_info = []
#     with zipfile.ZipFile(zip_path, 'r') as z:
#         if 'xl/connections.xml' in z.namelist():
#             with z.open('xl/connections.xml') as f:
#                 tree = ET.parse(f)
#                 root = tree.getroot()
#                 ns = {'default': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
#                 for conn in root.findall('default:connection', ns):
#                     conn_info = {k: conn.get(k) for k in conn.keys()}
#                     # Prova a estrarre dbPr, webPr, oledbPr, filePr, etc.
#                     dbPr = conn.find('default:dbPr', ns)
#                     webPr = conn.find('default:webPr', ns)
#                     oledbPr = conn.find('default:oledbPr', ns)
#                     filePr = conn.find('default:filePr', ns)
#                     if dbPr is not None:
#                         conn_info['dbPr'] = {k: dbPr.get(k) for k in dbPr.keys()}
#                     if webPr is not None:
#                         conn_info['webPr'] = {k: webPr.get(k) for k in webPr.keys()}
#                     if oledbPr is not None:
#                         conn_info['oledbPr'] = {k: oledbPr.get(k) for k in oledbPr.keys()}
#                     if filePr is not None:
#                         conn_info['filePr'] = {k: filePr.get(k) for k in filePr.keys()}
#                     connections_info.append(conn_info)
#     return connections_info

def proprieties_list(path):
    wb = load_workbook(path)
    props = wb.properties

    connection_sources = []
    with zipfile.ZipFile(path) as zF:
        if "xl/connections.xml" in zF.namelist():
            connections = extract_connections(path)
            if connections:
                for conn in connections:
                    # Mostra la sorgente (file/db/url) della connessione
                    source = conn.get('connection_string', 'N/A')
                    connection_sources.append(source)
            else:
                connection_sources.append('N/A')
            return [
                "File Name: " + os.path.basename(path),
                "Creato da: " + str(props.creator),
                "Modificato da: " + str(props.lastModifiedBy),
                "Nome Propriet\u00E0: " + str(props.title),
                "Data Creazione: " + str(props.created),
                "Data Modifica: " + str(props.modified),
                "Conessione esterna: Si",
                "Sorgenti connessione:",
                *[f"- {src}" for src in connection_sources]
            ]
        else:
            return [
                "File Name: " + os.path.basename(path),
                "Creato da: " + str(props.creator),
                "Modificato da: " + str(props.lastModifiedBy),
                "Nome Propriet\u00E0: " + str(props.title),
                "Data Creazione: " + str(props.created),
                "Data Modifica: " + str(props.modified),
                "Conessione esterna: No",
                "Sorgenti connessione: Nessuna"
            ]
        
# Funzione per estrarre query Power Query da xl/queries/*.xml
def extract_queries(zip_path):
    queries_info = []
    with zipfile.ZipFile(zip_path, 'r') as z:
        query_files = [f for f in z.namelist() if f.startswith('xl/queries/') and f.endswith('.xml')]
        for qf in query_files:
            with z.open(qf) as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'pq': 'http://schemas.microsoft.com/office/PowerQuery/2014/11/workbook'}
                name = root.get('name')
                formula_elem = root.find('pq:formula', ns)
                formula = formula_elem.text if formula_elem is not None else 'N/A'
                queries_info.append({
                    'file': qf,
                    'name': name,
                    'formula': formula
                })
    return queries_info


metadati = []
for file in files_name:
    metadati.append(proprieties_list(file))


# Stampa formattata dei metadati

for idx, file_meta in enumerate(metadati):
    print("="*60)
    for item in file_meta:
        print(item)
    # Estrai e mostra le query Power Query (sorgenti dati)
    file_path = files_name[idx]
    queries = extract_queries(file_path)
    if queries:
        print("Query Power Query trovate e sorgente dati originale:")
        for q in queries:
            print(f"- Nome: {q['name']}")
            print(f"  Formula completa:")
            print(f"  {q['formula']}")
            source_line = None
            for line in q['formula'].splitlines():
                if line.strip().lower().startswith('source ='):
                    source_line = line.strip()
                    break
            if source_line:
                print(f"  >>> Sorgente dati individuata: {source_line}")
            else:
                print("  Sorgenti dati non individuata")
        print("--- Tutte le formule M trovate sopra ---")
    else:
        print("Nessuna query Power Query trovata.")
    print()
print("="*60)
 
def extract_connections(zip_path):
    connections_info = []
    with zipfile.ZipFile(zip_path, 'r') as z:
        if 'xl/connections.xml' in z.namelist():
            with z.open('xl/connections.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'default': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for conn in root.findall('default:connection', ns):
                    conn_name = conn.get('name')
                    conn_type = conn.get('type')
                    conn_str = conn.find('default:dbPr', ns)
                    connection_string = conn_str.get('connection') if conn_str is not None else 'N/A'
                    connections_info.append({
                        'name': conn_name,
                        'type': conn_type,
                        'connection_string': connection_string
                    })
    return connections_info

def proprieties_list(path):
    wb = load_workbook(path)
    props = wb.properties

    connection_strings = []
    with zipfile.ZipFile(path) as zF:
        if "xl/connections.xml" in zF.namelist():
            # Estrai la stringa di connessione
            connections = extract_connections(path)
            if connections:
                for conn in connections:
                    connection_strings.append(conn['connection_string'])
            else:
                connection_strings.append('N/A')
            return [
                "File Name: " + os.path.basename(path),
                "Creato da: " + str(props.creator),
                "Modificato da: " + str(props.lastModifiedBy),
                "Nome Propriet\u00E0: " + str(props.title),
                "Data Creazione: " + str(props.created),
                "Data Modifica: " + str(props.modified),
                "Conessione esterna: Si",
                "Connection string: " + ", ".join(connection_strings)
            ]
        else:
            return [
                "File Name: " + os.path.basename(path),
                "Creato da: " + str(props.creator),
                "Modificato da: " + str(props.lastModifiedBy),
                "Nome Propriet\u00E0: " + str(props.title),
                "Data Creazione: " + str(props.created),
                "Data Modifica: " + str(props.modified),
                "Conessione esterna: No",
                "Connection string: "
            ]

# Funzione per estrarre query Power Query da xl/queries/*.xml

export_folder = r"C:\Users\giuseppe.tanda\Desktop\doValue\Export M Code"
# Percorso di export del file Excel finale
excel_export_path = r"C:\Users\giuseppe.tanda\Desktop\doValue\Connessioni Trovate.xlsx"

export_data = []
unique_keys = set()

def extract_source_from_formula(formula):
    """
    Estrae la sorgente (Source) e il nome della tabella ([Name] = ...) dal codice M di Power Query.
    Restituisce una tupla (sorgente, nome_tabella).
    """
    source = None
    sorgente = None
    table_name = None
    schema = None
    item = None
    # Regex per vari pattern di Name
    name_patterns = [
        r'\[Name\]\s*=\s*"([^"]+)"',           # [Name] = "..."
        r'Name\s*=\s*"([^"]+)"',                 # Name = "..."
        r'Name\s*:\s*"([^"]+)"',                 # Name: "..."
        r'\[Name="([^"]+)"\]',                   # [Name="..."]
        r'Library\s*=\s*Origine\{\[Name="([^"]+)"\}\}\[Data\]', # Library = Origine{[Name="..."]}[Data]
    ]
    schema_pattern = r'Schema\s*=\s*"([^"]+)"'
    item_pattern = r'Item\s*=\s*"([^"]+)"'
    lines = formula.splitlines()
    idx = 0
    while idx < len(lines):
        l = lines[idx].strip()
        if l.lower().startswith("source =") or l.lower().startswith("origine ="):
            source = l
            # Caso 1: Sql.Database("server", "db")
            db_match = re.search(r'Sql\.Database\(\s*[^,]+,\s*"([^"]+)"', l)
            if db_match:
                sorgente = db_match.group(1)
            # Caso 2: Sql.Databases("server") seguito da qualsiasi riga = Origine{[Name="..."]}[Data]
            elif re.search(r'Sql\.Databases\(', l):
                for j in range(idx+1, len(lines)):
                    next_l = lines[j].strip()
                    lib_match = re.search(r'=\s*Origine\{\[Name="([^"]+)"\}\}\[Data\]', next_l)
                    if lib_match:
                        sorgente = lib_match.group(1)
                        break
                # Se non trovato, fallback: cerca la prima riga dopo Sql.Databases che contiene Name="..."
                if not sorgente:
                    for j in range(idx+1, len(lines)):
                        next_l = lines[j].strip()
                        name_match = re.search(r'Name\s*=\s*"([^"]+)"', next_l)
                        if name_match:
                            sorgente = name_match.group(1)
                            break
        # Fallback: cerca Library = Origine{[Name="..."]}[Data] ovunque
        if not sorgente:
            lib_match = re.search(r'Library\s*=\s*Origine\{\[Name="([^"]+)"\}\}\[Data\]', l)
            if lib_match:
                sorgente = lib_match.group(1)
        for pat in name_patterns:
            match = re.search(pat, l)
            if match:
                table_name = match.group(1)
                break
        if not schema:
            match_schema = re.search(schema_pattern, l)
            if match_schema:
                schema = match_schema.group(1)
        if not item:
            match_item = re.search(item_pattern, l)
            if match_item:
                item = match_item.group(1)
        idx += 1
        for pat in name_patterns:
            match = re.search(pat, l)
            if match:
                table_name = match.group(1)
                break
        if not schema:
            match_schema = re.search(schema_pattern, l)
            if match_schema:
                schema = match_schema.group(1)
        if not item:
            match_item = re.search(item_pattern, l)
            if match_item:
                item = match_item.group(1)

    # Logica combinata: se sorgente SQL, usa nome DB, altrimenti usa la stringa source
    if sorgente:
        final_sorgente = sorgente
    elif source:
        final_sorgente = source
    else:
        final_sorgente = "Sorgente non individuata"
    if not table_name:
        table_name = "Tabella non individuata"
    if not schema:
        schema = "Schema non individuato"
    if not item:
        item = "Item non individuato"
    return source, final_sorgente, schema, item

for root, dirs, files in os.walk(export_folder):
    for txt_file in files:
        if txt_file.endswith(".txt"):
            txt_path = os.path.join(root, txt_file)
            file_name = txt_file.split('_')[0] + ".xlsx"
            folder_rel = os.path.relpath(root, export_folder)
            # Associa solo se il file Excel esiste nella stessa sottocartella relativa
            xlsx_path = os.path.join(folder_path, folder_rel, file_name)
            if os.path.exists(xlsx_path):
                connessione_si_no = "Si"
                sorgente = None
                table_name = None
                schema = None
                item = None
                kind = None
                creator = lastModifiedBy = title = created = modified = None
                try:
                    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                    props = wb.properties
                    creator = str(props.creator) if props.creator else None
                    lastModifiedBy = str(props.lastModifiedBy) if props.lastModifiedBy else None
                    title = str(props.title) if props.title else None
                    created = str(props.created) if props.created else None
                    modified = str(props.modified) if props.modified else None
                except Exception:
                    pass
                with open(txt_path, encoding="utf-8") as f:
                    formula = f.read()
                    source, final_sorgente, schema, item = extract_source_from_formula(formula)
                export_data.append({
                    "FileName": file_name,
                    "FolderName": folder_rel,
                    "Connessione": connessione_si_no,
                    "ConnessioneSorgente": source,
                    "Sorgente": final_sorgente,
                    "Schema": schema,
                    "TableName": item,
                    "Creator": creator,
                    "LastModifiedBy": lastModifiedBy,
                    "Created": created,
                    "Modified": modified
                })

# Trova anche i file senza connessione esportata
for file in files_name:
    folder_rel = os.path.relpath(os.path.dirname(file), folder_path)
    creator = lastModifiedBy = title = created = modified = None
    connessione = "No"
    sorgente = "Nessuna connessione Power Query esportata"
    table_name = "Tabella non individuata"
    schema = "Schema non individuato"
    item = "Item non individuato"
    kind = "Kind non individuato"
    file_name = os.path.basename(file)
    # Estrai metadati
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        props = wb.properties
        creator = str(props.creator) if props.creator else None
        lastModifiedBy = str(props.lastModifiedBy) if props.lastModifiedBy else None
        title = str(props.title) if props.title else None
        created = str(props.created) if props.created else None
        modified = str(props.modified) if props.modified else None
    except Exception as e:
        print(f"Errore nel file {file}: {e}")
    # Cerca file .txt associati SOLO nella sottocartella corrispondente
    mcode_found = False
    export_mcode_subfolder = os.path.join(export_folder, folder_rel)
    if os.path.exists(export_mcode_subfolder):
        for txt_file in os.listdir(export_mcode_subfolder):
            if txt_file.endswith(".txt") and txt_file.startswith(file_name.replace('.xlsx','')):
                txt_path = os.path.join(export_mcode_subfolder, txt_file)
                with open(txt_path, encoding="utf-8") as f:
                    formula = f.read()
                    source, final_sorgente, schema, item = extract_source_from_formula(formula)
                    connessione = "Si"
                    mcode_found = True
                    unique_key = f"{file_name}|{folder_rel}|{source}|{item}"
                    if unique_key not in unique_keys:
                        unique_keys.add(unique_key)
                        export_data.append({
                            "FileName": file_name,
                            "FolderName": folder_rel,
                            "Connessione": connessione,
                            "ConnessioneSorgente": source,
                            "Sorgente": final_sorgente,
                            "Schema": schema,
                            "TableName": item,
                            "Creator": creator,
                            "LastModifiedBy": lastModifiedBy,
                            "Created": created,
                            "Modified": modified
                        })
    # Se non è stato trovato nessun codice M, esporta solo i metadati
    if not mcode_found:
        unique_key = f"{file_name}|{folder_rel}|{sorgente}"
        if unique_key not in unique_keys:
            unique_keys.add(unique_key)
            export_data.append({
                "FileName": file_name,
                "FolderName": folder_rel,
                "Connessione": connessione,
                "ConnessioneSorgente": sorgente,
                "DatabaseName": table_name,
                "Schema": schema,
                "TableName": item,
                "Kind": kind,
                "Creator": creator,
                "LastModifiedBy": lastModifiedBy,
                "Created": created,
                "Modified": modified
            })

# Esporta in Excel
df = pd.DataFrame(export_data)
# Elimina DatabaseName e Kind se presenti
for col in ["DatabaseName", "Kind"]:
    if col in df.columns:
        df = df.drop(columns=[col])
df = df.drop_duplicates(subset=["FileName", "FolderName", "ConnessioneSorgente", "TableName"])
df.to_excel(excel_export_path, index=False)
print(f"\nFile Excel esportato: {excel_export_path}")
# -------------------------------------------------------------
# 1. SCRIPT POWERSHELL PER ESTRARRE IL CODICE M DA FILE EXCEL
# -------------------------------------------------------------
# Salva questo script come ExportMCode.ps1 e modificalo con i tuoi percorsi
# Esegui da PowerShell prima di lanciare lo script Python
#
# $excel = New-Object -ComObject Excel.Application
# $excel.Visible = $false
# $folder = "C:\Percorso\Cartella\Excel"  # <-- Modifica con la tua cartella
# $exportFolder = "C:\Percorso\Export\MCode"
# New-Item -ItemType Directory -Force -Path $exportFolder
# $files = Get-ChildItem -Path $folder -Filter *.xlsx
# foreach ($file in $files) {
#     $wb = $excel.Workbooks.Open($file.FullName)
#     foreach ($query in $wb.Queries) {
#         $queryName = $query.Name
#         $mCode = $query.Formula
#         $exportPath = Join-Path $exportFolder "$($file.BaseName)_$($queryName)_M.txt"
#         Set-Content -Path $exportPath -Value $mCode
#     }
#     $wb.Close($false)
# }
# $excel.Quit()
# -------------------------------------------------------------

# -------------------------------------------------------------
# 2. SCRIPT PYTHON PER INDIVIDUARE LA SORGENTE NEL CODICE M
# -------------------------------------------------------------
export_folder = r"C:\Users\giuseppe.tanda\Desktop\doValue\Export M Code"  # <-- Modifica con il tuo percorso

print("\nRISULTATI SORGENTI POWER QUERY DA FILE M ESPORTATI:\n")
for txt_file in os.listdir(export_folder):
    if txt_file.endswith(".txt"):
        print("="*60)
        print(f"File: {txt_file}")
        with open(os.path.join(export_folder, txt_file), encoding="utf-8") as f:
            lines = f.readlines()
            found = False
            for line in lines:
                if line.strip().lower().startswith("source ="):
                    print(f"Sorgente dati individuata: {line.strip()}")
                    found = True
            if not found:
                print("Sorgenti dati non individuata")
        print("="*60)


