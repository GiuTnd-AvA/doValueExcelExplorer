import os
import re
from typing import List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore
    Alignment = None  # type: ignore


class PowerQueryTxtSourceExtractor:
    """
    Scans a root directory recursively, finds .txt files that contain exported
    Power Query M code, extracts the first complete 'Source =' line, and can
    write results to an Excel report.
    """

    # Match any line containing 'Source =' (EN) or 'Origine =' (IT),
    # even if preceded by 'let' or other characters on the same line
    SOURCE_PATTERN = re.compile(r"^.*\b(Source|Origine)\s*=.*$", re.MULTILINE | re.IGNORECASE)

    def __init__(self, root_dir: str) -> None:
        self.root_dir = root_dir
        # rows: (full_path, file_name, source_line)
        self.rows: List[Tuple[str, str, str]] = []
        self.missing_files: List[str] = []     # relative paths of .txt without 'Source ='

    @staticmethod
    def _read_text_best_effort(path: str) -> Optional[str]:
        """Read text with common encodings, returning None on failure."""
        for enc in ("utf-8", "utf-16", "latin-1", "cp1252"):
            try:
                with open(path, "r", encoding=enc, errors="strict") as f:
                    return f.read()
            except Exception:
                continue
        return None

    @classmethod
    def extract_source_line(cls, content: str) -> Optional[str]:
        """Return the first line containing 'Source =' or 'Origine =' from content, if any."""
        m = cls.SOURCE_PATTERN.search(content)
        if not m:
            return None
        line = m.group(0).strip()
        return line

    def scan(self, verbose: bool = True) -> List[Tuple[str, str, str]]:
        """Populate self.rows with (file_name, source_line) for each .txt file.
        Returns the rows list.
        """
        self.rows.clear()
        self.missing_files.clear()

        # First pass: count total .txt files
        total_txt = 0
        for dirpath, _, filenames in os.walk(self.root_dir):
            total_txt += sum(1 for fn in filenames if fn.lower().endswith(".txt"))
        if verbose:
            print(f"Total .txt files to process: {total_txt}")

        processed = 0
        found = 0
        for dirpath, _, filenames in os.walk(self.root_dir):
            if verbose:
                print(f"Scanning folder: {dirpath}")
            for fname in filenames:
                if not fname.lower().endswith(".txt"):
                    continue
                processed += 1
                full_path = os.path.join(dirpath, fname)
                if verbose:
                    print(f"[{processed}/{total_txt}] Processing: {fname}")

                content = self._read_text_best_effort(full_path)
                if content is None:
                    if verbose:
                        print(f"WARN: Could not read file: {full_path}")
                    continue

                src_line = self.extract_source_line(content)
                if src_line:
                    self.rows.append((full_path, fname, src_line))
                    found += 1
                else:
                    # Store relative path of files missing 'Source ='
                    rel = os.path.relpath(full_path, self.root_dir)
                    self.missing_files.append(rel)

        if verbose:
            print(f"Scan complete: processed {processed}/{total_txt} .txt files, found {found} with 'Source/Origine ='.")
            if self.missing_files:
                print(f"Files without 'Source/Origine =': {len(self.missing_files)}")
                for mf in self.missing_files:
                    print(f" - {mf}")
            else:
                print("All processed .txt files contain a 'Source/Origine =' line.")
        return self.rows

    def write_report(self, output_path: str) -> str:
        if Workbook is None:
            raise RuntimeError("openpyxl not installed; cannot write Excel report.")

        # Prepare rows with safe truncation for Excel cell limit
        max_cell_len = 32767
        headers = ["Path", "File", "Source"]
        safe_rows = []
        for full_path, fname, source in self.rows:
            val = source if len(source) <= max_cell_len else (source[: max_cell_len - 15] + "... [TRUNCATED]")
            safe_rows.append((full_path, fname, val))

        widths = [80, 40, 120]
        try:
            from Report.Excel_Writer import write_rows_split_across_files
        except Exception:
            write_rows_split_across_files = None  # type: ignore

        if write_rows_split_across_files is not None:
            write_rows_split_across_files(headers, safe_rows, output_path, sheet_name="PowerQuery Sources", column_widths=widths)
            return os.path.abspath(output_path)
        else:
            # Fallback single-file
            wb = Workbook()
            ws = wb.active
            ws.title = "PowerQuery Sources"
            ws.append(headers)
            for r in safe_rows:
                ws.append(list(r))
            from openpyxl.utils import get_column_letter
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            # Ensure directory exists
            out_dir = os.path.dirname(output_path)
            if out_dir and not os.path.exists(out_dir):
                os.makedirs(out_dir, exist_ok=True)
            wb.save(output_path)
            return os.path.abspath(output_path)
