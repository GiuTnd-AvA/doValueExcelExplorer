import argparse
import os
from typing import List, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from BusinessLogic.PowerQuerySourceConnectionParser import PowerQuerySourceConnectionParser


def read_sources_from_excel(input_path: str) -> List[Tuple[str, str, str]]:
    """
    Returns a list of tuples (path, file, source_line) from the report produced by Export_PowerQuery_Sources.py
    Expected headers: Path | File | Source
    """
    wb = load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {h: i for i, h in enumerate(headers)}
    required = ["Path", "File", "Source"]
    for r in required:
        if r not in header_index:
            raise ValueError(f"Input Excel missing required column '{r}'. Found: {headers}")

    rows: List[Tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        path = row[header_index["Path"]]
        file = row[header_index["File"]]
        source = row[header_index["Source"]]
        if source:
            rows.append((path or "", file or "", source))
    return rows


def write_parsed_excel(entries: List[Tuple[str, str, str, str, str, str, str, str]], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ParsedConnections"

    # Include file path, original Source, and aggregated Join tables
    headers = ["Path", "File", "Server", "Database", "Schema", "Table", "Join", "Source"]
    ws.append(headers)

    for path, file, server, database, schema, table, join_str, source in entries:
        ws.append([path, file, server, database, schema, table, join_str, source])

    # Simple sizing
    widths = [60, 28, 20, 20, 16, 40, 60, 100]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser(description="Parse PowerQuery Source lines into connection parts")
    ap.add_argument("--in", dest="input", default=os.path.join("Report", "PowerQuery_Sources_Report.xlsx"),
                    help="Path to the input Excel with Source lines (default: Report/PowerQuery_Sources_Report.xlsx)")
    ap.add_argument("--out", dest="output", default=os.path.join("Report", "PowerQuery_Parsed_Connections.xlsx"),
                    help="Output path for the parsed connections Excel")
    args = ap.parse_args()

    rows = read_sources_from_excel(args.input)
    parser = PowerQuerySourceConnectionParser()

    parsed: List[Tuple[str, str, str, str, str, str, str, str]] = []
    for path, file, source in rows:
        infos = parser.parse_all(source)
        if not infos:
            # Fall back to single parse for robustness
            info = parser.parse(source)
            infos = [info] if info else []

        if not infos:
            # still nothing; output minimal row
            parsed.append((path, file, None, None, None, None, "", source))
            continue

        primary = infos[0]
        others = infos[1:]

        def fmt(i: dict) -> str:
            parts = [i.get("database"), i.get("schema"), i.get("table")]
            return ".".join([p for p in parts if p])

        join_str = "; ".join([fmt(i) for i in others])

        parsed.append((
            path,
            file,
            primary.get("server"),
            primary.get("database"),
            primary.get("schema"),
            primary.get("table"),
            join_str,
            source,
        ))

    write_parsed_excel(parsed, args.output)
    print(f"Parsed {len(parsed)} entries. Report written: {os.path.abspath(args.output)}")


if __name__ == "__main__":
    main()
