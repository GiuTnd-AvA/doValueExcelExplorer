# -------------------------------------------------------------------
# Scopo: prende un file .sql e lo legge per estrarre tutte le tabelle 
# e la clausola con cui viene chiamata
# -------------------------------------------------------------------

import argparse
import os
import re
import csv
import datetime
from typing import List, Dict, Tuple

try:
    import openpyxl
except ImportError:  # Fallback if not installed
    openpyxl = None

# User-configurable default input path. If you don't pass the positional
# argument, the script will use this path.
# Update this value to your .sql file path.
INPUT_SQL = r"c:/Users/giuseppe.tanda/Desktop/doValue/OneDrive_1_12-30-2025/Append SQL script.sql"

# Identifier: [name] or "name" or `name` or simple names, including temp tables #/##
IDENTIFIER = r'(?:\[[^\]]+\]|"[^"]+"|`[^`]+`|[#]{1,2}[A-Za-z_][A-Za-z0-9_$]*|[A-Za-z_][A-Za-z0-9_$]*)'
# Qualified names support:
# - table
# - schema.table
# - db.schema.table
# - db..table   (missing schema -> default to dbo)
# The regex below allows either 2-part names or 3-part names where the middle part may be empty
QUALIFIED = rf'{IDENTIFIER}(?:\s*\.\s*(?:{IDENTIFIER}|(?=\s*\.))\s*\.\s*{IDENTIFIER}|\s*\.\s*{IDENTIFIER})?'

# Detect pattern db..table to normalize missing schema to dbo
EMPTY_SCHEMA_PATTERN = re.compile(rf'^(?P<db>{IDENTIFIER})\s*\.\s*\.\s*(?P<table>{IDENTIFIER})$', re.IGNORECASE)

CLAUSE_PATTERNS: List[Tuple[str, re.Pattern]] = [
    ("DROP TABLE",     re.compile(rf"\bDROP\s+TABLE\s+(?:IF\s+EXISTS\s+)?(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("TRUNCATE TABLE", re.compile(rf"\bTRUNCATE\s+TABLE\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("CREATE TABLE",   re.compile(rf"\bCREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("ALTER TABLE",    re.compile(rf"\bALTER\s+TABLE\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("INSERT INTO",    re.compile(rf"\bINSERT\s+INTO\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("SELECT INTO",    re.compile(rf"\bSELECT\b[\s\S]*?\bINTO\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("UPDATE",         re.compile(rf"\bUPDATE\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("DELETE FROM",    re.compile(rf"\bDELETE\s+FROM\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    ("MERGE INTO",     re.compile(rf"\bMERGE\s+INTO\s+(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    # FROM: allow optional parenthesis when not starting a subquery (SELECT/WITH)
    ("FROM",           re.compile(rf"\bFROM\s+(?:\(\s*(?!SELECT\b|WITH\b))?(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    # JOIN: similarly allow optional parenthesis when not a subquery
    ("JOIN",           re.compile(rf"(?:(?P<type>INNER|LEFT|RIGHT|FULL(?:\s+OUTER)?|CROSS)\s+)?JOIN\s+(?:\(\s*(?!SELECT\b|WITH\b))?(?P<table>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
]

MARKER_REGEX = re.compile(r"(?m)^\s*--\s*(?P<num>\d+)\s+(?P<path>.+?\.sql)\s*$")


def extract_matches(text: str) -> List[Dict[str, str]]:
    # Collect all matches with the position of the TABLE token,
    # then sort by that position to reflect exact encounter order.
    collected: List[Dict[str, str]] = []
    for clause, pattern in CLAUSE_PATTERNS:
        for m in pattern.finditer(text):
            c = clause
            t = m.group('table').strip()
            # Normalize db..table -> db.dbo.table
            em = EMPTY_SCHEMA_PATTERN.match(t)
            if em:
                t = f"{em.group('db')}.dbo.{em.group('table')}"
            join_type = m.groupdict().get('type')
            if join_type:
                c = f"{join_type.upper()} JOIN"
            # Prefer the position where the table token appears
            try:
                pos = m.start('table')
            except Exception:
                pos = m.start()
            collected.append({'Clause': c, 'Table': t, '_pos': pos})
    collected.sort(key=lambda x: x['_pos'])
    # Drop position before returning
    return [{'Clause': x['Clause'], 'Table': x['Table']} for x in collected]


def parse_blocks(content: str, input_path: str, verbose: bool = True) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    markers = list(MARKER_REGEX.finditer(content))
    if not markers:
        # Single block case
        if verbose:
            print("Nessun marker trovato; elaboro il file come singolo blocco…")
        file_name = os.path.basename(input_path)
        matches = extract_matches(content)
        if verbose:
            print(f"  Riferimenti estratti: {len(matches)}")
        for m in matches:
            rows.append({'Path': input_path, 'File': file_name, 'Clause': m['Clause'], 'Table': m['Table']})
        return rows

    # Multiple blocks
    if verbose:
        print(f"Trovati {len(markers)} marker di file; elaboro i blocchi…")
    for i, marker in enumerate(markers):
        path = marker.group('path').strip()
        file_name = os.path.basename(path)
        start = marker.end()
        end = markers[i+1].start() if i < len(markers)-1 else len(content)
        block = content[start:end]
        matches = extract_matches(block)
        if verbose:
            print(f"  Blocco {i+1}/{len(markers)}: {file_name}")
            print(f"    Riferimenti estratti: {len(matches)}")
        for m in matches:
            rows.append({'Path': path, 'File': file_name, 'Clause': m['Clause'], 'Table': m['Table']})
    return rows


def write_csv(rows: List[Dict[str, str]], output_path: str) -> None:
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Path', 'File', 'Clause', 'Table'])
        for r in rows:
            w.writerow([r['Path'], r['File'], r['Clause'], r['Table']])


def write_xlsx(rows: List[Dict[str, str]], output_path: str) -> None:
    if openpyxl is None:
        raise RuntimeError('openpyxl is not installed')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tables'

    # Headers
    ws.cell(row=1, column=1, value='Path')
    ws.cell(row=1, column=2, value='File')
    ws.cell(row=1, column=3, value='Clause')
    ws.cell(row=1, column=4, value='Table')

    for idx, r in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=r['Path'])
        ws.cell(row=idx, column=2, value=r['File'])
        ws.cell(row=idx, column=3, value=r['Clause'])
        ws.cell(row=idx, column=4, value=r['Table'])

    # Autofit (approx by setting column width to max length)
    for col in range(1, 5):
        max_len = 0
        for row in range(1, len(rows)+2):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 80)

    try:
        wb.save(output_path)
    except PermissionError:
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        base, ext = os.path.splitext(output_path)
        alt_path = f"{base}_{ts}{ext}"
        print(f"File Excel destinazione bloccato: salvo come {alt_path}")
        wb.save(alt_path)


def main():
    ap = argparse.ArgumentParser(description='Extract table references from appended SQL into Excel/CSV.')
    ap.add_argument('input', nargs='?', help='Path to appended SQL file (optional if INPUT_SQL is set)')
    ap.add_argument('-o', '--output', help='Output file path (.xlsx or .csv)')
    ap.add_argument('-f', '--format', choices=['xlsx', 'csv'], default='xlsx', help='Output format')
    args = ap.parse_args()

    # Resolve input path: CLI arg takes precedence; otherwise use INPUT_SQL
    input_path = args.input if args.input else INPUT_SQL
    if not input_path:
        raise FileNotFoundError('Percorso input non fornito: passa l\'argomento "input" oppure imposta INPUT_SQL all\'inizio dello script.')
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'Input file not found: {input_path}')

    # Default output path
    if not args.output:
        ext = 'xlsx' if args.format == 'xlsx' else 'csv'
        args.output = os.path.join(os.path.dirname(input_path), f'SQL_Table_References.{ext}')

    print('Avvio estrazione…')
    print(f'Input SQL: {input_path}')
    with open(input_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    rows = parse_blocks(content, input_path, verbose=True)

    # Preserve original encounter order (no sorting)
    print(f'Totale riferimenti trovati: {len(rows)}')

    if args.format == 'csv':
        print(f'Scrittura CSV in {args.output}…')
        write_csv(rows, args.output)
        print(f'CSV written to {args.output}')
    else:
        print(f'Scrittura Excel in {args.output}…')
        write_xlsx(rows, args.output)
        print(f'Excel written to {args.output}')


if __name__ == '__main__':
    main()
