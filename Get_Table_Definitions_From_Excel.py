# -----------------------------------------------------------------------------
# Scopo: legge un file con gli elementi di connessione e per ogni elementi mi dice
# il tipo di oggetto di DB che sta leggendo e nel database cerca la devinizione 
# -----------------------------------------------------------------------------

import os
from typing import List, Optional, Tuple, Dict

try:
    import pyodbc
except Exception:
    pyodbc = None  # type: ignore

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore

try:
    import pandas as pd
except Exception:
    pd = None  # type: ignore

# -----------------------------------------------------------------------------
# Config: imposta i percorsi di input e output (Excel)
# -----------------------------------------------------------------------------
INPUT_EXCEL_PATH: Optional[str] = None   # es: r"C:\\path\\liste_tabelle.xlsx"
OUTPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\ddl_tabelle.xlsx"

# Default connessione
DEFAULT_SERVER: str = "EPCP3"
DEFAULT_DB: str = "master"

ODBC_DRIVERS: List[str] = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
]
TRUSTED_CONNECTION: bool = True
SQL_USERNAME: Optional[str] = None
SQL_PASSWORD: Optional[str] = None
CONNECTION_TEST_TIMEOUT: int = 3
QUERY_TIMEOUT: int = 60
# Regola se il server richiede Encrypt
ODBC_ENCRYPT_OPTS: str = "Encrypt=no;TrustServerCertificate=yes;"


class TableDefinitionExtractor:
    """Legge la lista tabelle da Excel (colonne: Server | DB | Schema | Table)
    e produce un Excel di output con le stesse colonne + DDL.
    """

    def __init__(self, input_excel: str, output_excel: str):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if pd is None:
            raise RuntimeError("pandas non installato. Installa 'pip install pandas openpyxl'.")
        if pyodbc is None:
            raise RuntimeError("pyodbc non installato. Installa 'pip install pyodbc'.")
        if not input_excel:
            raise ValueError("Imposta INPUT_EXCEL_PATH.")
        self.input_excel = input_excel
        self.output_excel = output_excel or os.path.join(os.path.dirname(input_excel) or os.getcwd(), "DDL_Tabelle.xlsx")

    # ---------------- Excel parsing ----------------
    def _read_items(self) -> List[Tuple[str, str, str, str]]:
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not rows:
                return []

            headers = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            # Serve almeno: server, db|database, schema, table
            required = {"server", "schema", "table"}
            has_db = ("db" in headers) or ("database" in headers)
            if not (required.issubset(set(headers)) and has_db):
                raise RuntimeError("Il foglio deve contenere le colonne: Server, DB (o Database), Schema, Table.")

            idx_server = headers.index("server")
            idx_db = headers.index("db") if "db" in headers else headers.index("database")
            idx_schema = headers.index("schema")
            idx_table = headers.index("table")

            items: List[Tuple[str, str, str, str]] = []
            for r in rows[1:]:
                if r is None:
                    continue
                server = str(r[idx_server]).strip() if r[idx_server] else DEFAULT_SERVER
                db = str(r[idx_db]).strip() if r[idx_db] else DEFAULT_DB
                schema = str(r[idx_schema]).strip() if r[idx_schema] else "dbo"
                table = str(r[idx_table]).strip() if r[idx_table] else ""
                if not table:
                    continue
                items.append((server, db, schema, table))
            return items
        finally:
            wb.close()

    # ---------------- Connection handling ----------------
    def _build_conn_str(self, server: str, db: str) -> str:
        last_error: Optional[Exception] = None
        for drv in ODBC_DRIVERS:
            try:
                conn_str = f"DRIVER={{{drv}}};SERVER={server};DATABASE={db};"
                if TRUSTED_CONNECTION:
                    conn_str += "Trusted_Connection=yes;"
                else:
                    if not SQL_USERNAME or not SQL_PASSWORD:
                        raise RuntimeError("Credenziali non impostate e Trusted_Connection disattivata.")
                    conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                conn_str += ODBC_ENCRYPT_OPTS
                # test driver
                conn = pyodbc.connect(conn_str, timeout=CONNECTION_TEST_TIMEOUT)
                conn.close()
                return conn_str
            except Exception as e:
                last_error = e
                continue
        raise RuntimeError(f"Nessun driver ODBC valido trovato per {server}/{db}. Ultimo errore: {last_error}")

    # ---------------- DDL builder ----------------
    def _fetch_table_ddl(self, conn, schema: str, table: str) -> str:
        tsql_stringagg = r"""
DECLARE @schema_table nvarchar(512) = QUOTENAME(?) + N'.' + QUOTENAME(?);
DECLARE @obj_id int = OBJECT_ID(@schema_table);
IF @obj_id IS NULL
    SELECT CAST(N'ERROR: tabella non trovata: ' + @schema_table AS nvarchar(max)) AS ddl;
ELSE
BEGIN
    DECLARE @cols nvarchar(max) =
    (
        SELECT STRING_AGG(
            N'[' + c.name + N'] ' +
            UPPER(t.name) +
            CASE 
                WHEN t.name IN (N'char',N'nchar',N'varchar',N'nvarchar',N'binary',N'varbinary') 
                     THEN N'(' + CASE 
                                   WHEN t.name IN (N'nchar',N'nvarchar') 
                                        THEN CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length/2 AS nvarchar(10)) END
                                   ELSE CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length AS nvarchar(10)) END
                                 END + N')'
                WHEN t.name IN (N'decimal',N'numeric') 
                     THEN N'(' + CAST(c.precision AS nvarchar(10)) + N',' + CAST(c.scale AS nvarchar(10)) + N')'
                WHEN t.name IN (N'datetime2',N'time',N'datetimeoffset')
                     THEN N'(' + CAST(c.scale AS nvarchar(10)) + N')'
                ELSE N''
            END +
            CASE WHEN ic.is_identity = 1 
                 THEN N' IDENTITY(' + CAST(ic.seed_value AS nvarchar(50)) + N',' + CAST(ic.increment_value AS nvarchar(50)) + N')' 
                 ELSE N'' 
            END +
            CASE WHEN c.is_nullable = 0 THEN N' NOT NULL' ELSE N' NULL' END +
            COALESCE(N' DEFAULT ' + dc.definition, N'')
            , N',' + CHAR(13) + CHAR(10)
        ) WITHIN GROUP (ORDER BY c.column_id)
        FROM sys.columns c
        JOIN sys.types t ON c.user_type_id = t.user_type_id
        LEFT JOIN sys.default_constraints dc ON dc.parent_object_id = c.object_id AND dc.parent_column_id = c.column_id
        LEFT JOIN sys.identity_columns ic ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        WHERE c.object_id = @obj_id
    );

    DECLARE @pk nvarchar(max) =
    (
        SELECT N'CONSTRAINT [' + kc.name + N'] PRIMARY KEY (' +
               STRING_AGG(N'[' + c.name + N']' + CASE WHEN ic.is_descending_key = 1 THEN N' DESC' ELSE N'' END, N', ')
               WITHIN GROUP (ORDER BY ic.key_ordinal) + N')'
        FROM sys.key_constraints kc
        JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
        JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id
        JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
        WHERE kc.type = 'PK' AND kc.parent_object_id = @obj_id
    );

    SELECT N'CREATE TABLE ' + @schema_table + CHAR(13) + CHAR(10) +
           N'(' + ISNULL(@cols, N'') + N')' +
           ISNULL(CHAR(13) + CHAR(10) + @pk, N'') AS ddl;
END
"""
        tsql_xmlpath = r"""
DECLARE @schema_table nvarchar(512) = QUOTENAME(?) + N'.' + QUOTENAME(?);
DECLARE @obj_id int = OBJECT_ID(@schema_table);
IF @obj_id IS NULL
    SELECT CAST(N'ERROR: tabella non trovata: ' + @schema_table AS nvarchar(max)) AS ddl;
ELSE
BEGIN
    DECLARE @cols nvarchar(max) = N'';
    SELECT @cols = STUFF((
        SELECT N',' + CHAR(13) + CHAR(10) +
               N'[' + c.name + N'] ' +
               UPPER(t.name) +
               CASE 
                   WHEN t.name IN (N'char',N'nchar',N'varchar',N'nvarchar',N'binary',N'varbinary') 
                        THEN N'(' + CASE 
                                      WHEN t.name IN (N'nchar',N'nvarchar') 
                                           THEN CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length/2 AS nvarchar(10)) END
                                      ELSE CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length AS nvarchar(10)) END
                                    END + N')'
                   WHEN t.name IN (N'decimal',N'numeric') 
                        THEN N'(' + CAST(c.precision AS nvarchar(10)) + N',' + CAST(c.scale AS nvarchar(10)) + N')'
                   WHEN t.name IN (N'datetime2',N'time',N'datetimeoffset')
                        THEN N'(' + CAST(c.scale AS nvarchar(10)) + N')'
                   ELSE N''
               END +
               CASE WHEN ic.is_identity = 1 
                    THEN N' IDENTITY(' + CAST(ic.seed_value AS nvarchar(50)) + N',' + CAST(ic.increment_value AS nvarchar(50)) + N')' 
                    ELSE N'' 
               END +
               CASE WHEN c.is_nullable = 0 THEN N' NOT NULL' ELSE N' NULL' END +
               COALESCE(N' DEFAULT ' + dc.definition, N'')
        FROM sys.columns c
        JOIN sys.types t ON c.user_type_id = t.user_type_id
        LEFT JOIN sys.default_constraints dc ON dc.parent_object_id = c.object_id AND dc.parent_column_id = c.column_id
        LEFT JOIN sys.identity_columns ic ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        WHERE c.object_id = @obj_id
        ORDER BY c.column_id
        FOR XML PATH(''), TYPE
    ).value('.', 'nvarchar(max)'), 1, 1, N'');

    DECLARE @pk nvarchar(max) = NULL;
    SELECT @pk = N'CONSTRAINT [' + kc.name + N'] PRIMARY KEY (' +
                 STUFF((
                    SELECT N', ' + N'[' + c.name + N']' + CASE WHEN ic.is_descending_key = 1 THEN N' DESC' ELSE N'' END
                    FROM sys.index_columns ic
                    JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                    WHERE ic.object_id = i.object_id AND ic.index_id = i.index_id
                    ORDER BY ic.key_ordinal
                    FOR XML PATH(''), TYPE
                 ).value('.', 'nvarchar(max)'), 1, 2, N'') + N')'
    FROM sys.key_constraints kc
    JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
    WHERE kc.type = 'PK' AND kc.parent_object_id = @obj_id;

    SELECT N'CREATE TABLE ' + @schema_table + CHAR(13) + CHAR(10) +
           N'(' + ISNULL(@cols, N'') + N')' +
           ISNULL(CHAR(13) + CHAR(10) + @pk, N'') AS ddl;
END
"""
        cur = conn.cursor()
        try:
            cur.execute(tsql_stringagg, (schema, table))
        except Exception:
            cur.execute(tsql_xmlpath, (schema, table))
        row = cur.fetchone()
        return row[0] if row else ""

    def _get_object_type(self, conn, schema: str, name: str) -> str:
        """Ritorna il tipo oggetto in forma leggibile (es. 'Tabella', 'Vista').
        Se non trovato, ritorna 'Non trovato'.
        """
        sql = (
            """
            SELECT type, type_desc
            FROM sys.objects
            WHERE object_id = OBJECT_ID(QUOTENAME(?) + '.' + QUOTENAME(?));
            """
        )
        cur = conn.cursor()
        try:
            cur.execute(sql, (schema, name))
            r = cur.fetchone()
            if not r:
                return "Non trovato"
            code = str(r[0]) if r[0] is not None else ""
            desc = str(r[1]) if r[1] is not None else ""
            mapping = {
                "U": "Tabella",
                "V": "Vista",
                "P": "Stored Procedure",
                "FN": "Funzione (scalar)",
                "IF": "Funzione (inline table)",
                "TF": "Funzione (table)",
                "TR": "Trigger",
                "S": "Tabella di sistema",
                "SN": "Synonym",
                "SO": "Sequence/Service Object",
            }
            return mapping.get(code, desc or code or "Sconosciuto")
        except Exception:
            return "Sconosciuto"

    # ---------------- Main ----------------
    def run(self) -> str:
        items = self._read_items()
        if not items:
            raise RuntimeError("Nessuna tabella valida trovata nell'Excel di input.")

        total = len(items)
        print(f"[DDL] Totale tabelle da elaborare: {total}")
        results: List[List[str]] = []
        conns: Dict[Tuple[str, str], pyodbc.Connection] = {}
        try:
            for idx, (server, db, schema, table) in enumerate(items, start=1):
                print(f"[DDL] Elaborazione {idx}/{total}: {server}.{db}.{schema}.{table}")
                key = (server, db)
                if key not in conns:
                    conn_str = self._build_conn_str(server, db)
                    conns[key] = pyodbc.connect(conn_str, timeout=QUERY_TIMEOUT)
                obj_type = self._get_object_type(conns[key], schema, table)
                ddl = self._fetch_table_ddl(conns[key], schema, table)
                results.append([server, db, schema, table, obj_type, ddl])
        finally:
            for c in conns.values():
                try:
                    c.close()
                except Exception:
                    pass

        # Write output
        out_dir = os.path.dirname(self.output_excel)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        df = pd.DataFrame(results, columns=["Server", "DB", "Schema", "Table", "ObjectType", "DDL"])
        with pd.ExcelWriter(self.output_excel, engine="openpyxl", mode="w") as w:
            df.to_excel(w, index=False, sheet_name="DDL")
        return self.output_excel


if __name__ == "__main__":
    if not INPUT_EXCEL_PATH:
        raise SystemExit("Imposta INPUT_EXCEL_PATH a inizio file.")
    extractor = TableDefinitionExtractor(INPUT_EXCEL_PATH, OUTPUT_EXCEL_PATH or "")
    out = extractor.run()
    print(f"DDL scritto in: {out}")
