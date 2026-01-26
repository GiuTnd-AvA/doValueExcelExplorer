import os
import sys
from typing import List

import openpyxl
from openpyxl import Workbook

# Ensure workspace root is in path
ROOT = os.path.dirname(os.path.dirname(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from estrazione_sp.Analyze_SQL_Clause_Usage import SQLClauseAnalyzer


def _write_input_excel(path: str, rows: List[List[object]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append([
        "server",
        "database",
        "schema",
        "table",
        "nome oggetto",
        "tipo oggetto",
        "script creazione",
    ])

    for r in rows:
        ws.append(r)

    wb.save(path)


def _read_output_rows(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        out.append(row)
    wb.close()
    return out


def test_finds_from_and_join_clauses(tmp_path):
    xlsx_in = tmp_path / "in.xlsx"
    xlsx_out = tmp_path / "out.xlsx"

    script = """
    CREATE PROCEDURE dbo.usp_Test
    AS
    SELECT *
    FROM dbo.MyTable t
    INNER JOIN dbo.OtherTable o ON o.Id = t.Id;
    """

    _write_input_excel(
        str(xlsx_in),
        rows=[["EPCP3", "db1", "dbo", "MyTable", "dbo.usp_Test", "Stored Procedure", script]],
    )

    SQLClauseAnalyzer(str(xlsx_in), str(xlsx_out), sheet_name="Sheet1").process()
    assert os.path.exists(xlsx_out)

    rows = _read_output_rows(str(xlsx_out))
    assert len(rows) == 1
    *_, clause = rows[0]

    # Clause should include FROM for dbo.MyTable (and not match OtherTable)
    assert "FROM" in clause


def test_finds_update_insert_delete_merge_alter_create_drop_truncate_select_into(tmp_path):
    xlsx_in = tmp_path / "in.xlsx"
    xlsx_out = tmp_path / "out.xlsx"

    scripts = {
        "UPDATE": "UPDATE dbo.MyTable SET c = 1 WHERE id = 1;",
        "INSERT INTO": "INSERT INTO dbo.MyTable(c) SELECT 1;",
        "DELETE FROM": "DELETE FROM dbo.MyTable WHERE id = 1;",
        "MERGE INTO": "MERGE INTO dbo.MyTable AS t USING dbo.Src AS s ON 1=0 WHEN MATCHED THEN UPDATE SET t.c=1;",
        "ALTER TABLE": "ALTER TABLE dbo.MyTable ADD c2 int;",
        "CREATE TABLE": "CREATE TABLE dbo.MyTable(id int);",
        "DROP TABLE": "DROP TABLE dbo.MyTable;",
        "TRUNCATE TABLE": "TRUNCATE TABLE dbo.MyTable;",
        "SELECT INTO": "SELECT 1 AS c INTO dbo.MyTable;",
    }

    rows = []
    for idx, (expected, script) in enumerate(scripts.items(), start=1):
        rows.append(["EPCP3", "db1", "dbo", "MyTable", f"obj{idx}", "", script])

    _write_input_excel(str(xlsx_in), rows=rows)

    SQLClauseAnalyzer(str(xlsx_in), str(xlsx_out), sheet_name="Sheet1").process()
    out_rows = _read_output_rows(str(xlsx_out))
    assert len(out_rows) == len(scripts)

    # Map by object name (col E)
    by_obj = {r[4]: r for r in out_rows}

    for idx, expected in enumerate(scripts.keys(), start=1):
        clause = by_obj[f"obj{idx}"][-1]
        assert expected in clause


def test_ignores_comments(tmp_path):
    xlsx_in = tmp_path / "in.xlsx"
    xlsx_out = tmp_path / "out.xlsx"

    script = """
    CREATE PROCEDURE dbo.usp_CommentOnly
    AS
    -- FROM dbo.MyTable
    /*
      INNER JOIN dbo.MyTable
    */
    SELECT 1;
    """

    _write_input_excel(
        str(xlsx_in),
        rows=[["EPCP3", "db1", "dbo", "MyTable", "dbo.usp_CommentOnly", "Stored Procedure", script]],
    )

    SQLClauseAnalyzer(str(xlsx_in), str(xlsx_out), sheet_name="Sheet1").process()
    rows = _read_output_rows(str(xlsx_out))
    assert rows[0][-1] == "Non trovata"


def test_matches_bracketed_and_fully_qualified_names(tmp_path):
    xlsx_in = tmp_path / "in.xlsx"
    xlsx_out = tmp_path / "out.xlsx"

    script = """
    CREATE PROCEDURE dbo.usp_Qualified
    AS
    SELECT * FROM [dbo].[MyTable];
    SELECT * FROM EPCP3.db1.dbo.MyTable;
    """

    _write_input_excel(
        str(xlsx_in),
        rows=[["EPCP3", "db1", "dbo", "MyTable", "dbo.usp_Qualified", "Stored Procedure", script]],
    )

    SQLClauseAnalyzer(str(xlsx_in), str(xlsx_out), sheet_name="Sheet1").process()
    rows = _read_output_rows(str(xlsx_out))

    clause = rows[0][-1]
    assert "FROM" in clause
