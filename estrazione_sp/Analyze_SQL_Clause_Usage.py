# -----------------------------------------------------------------------------
# Scopo: analizza un file Excel contenente oggetti SQL (stored procedure, trigger,
# function) e identifica come le tabelle vengono utilizzate all'interno degli script.
# Per ogni occorrenza della tabella nello script, estrae la SQL clause utilizzata
# (FROM, JOIN, INSERT INTO, UPDATE, DELETE, ALTER, etc.).
# Produce un file Excel con le stesse colonne di input più una colonna aggiuntiva
# che descrive la SQL clause con cui la tabella viene referenziata.
# -----------------------------------------------------------------------------

import os
import re
from typing import List, Optional, Tuple, Set

try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore

# -----------------------------------------------------------------------------
# CONFIGURAZIONE: inserisci qui i percorsi dei file
# -----------------------------------------------------------------------------

# Percorso del file Excel di input con gli oggetti SQL
INPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\path\to\oggetti_sql.xlsx"

# Percorso del file Excel di output con le SQL clause identificate
OUTPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\path\to\oggetti_sql_con_clause.xlsx"

# Nome del foglio nel file Excel di input (None = primo foglio)
INPUT_SHEET_NAME: Optional[str] = None


class SQLClauseAnalyzer:
    """Analizza script SQL per identificare come le tabelle vengono utilizzate."""

    # Pattern per identificare SQL identifier (con o senza delimitatori)
    IDENTIFIER = r'(?:\[[^\]]+\]|"[^"]+"|`[^`]+`|[#]{1,2}[A-Za-z_][A-Za-z0-9_$]*|[A-Za-z_][A-Za-z0-9_$]*)'
    
    # Pattern per nomi qualificati (server.db.schema.table, db.schema.table, schema.table, table)
    # Supporta anche db..table (schema vuoto)
    QUALIFIED_TABLE = rf'{IDENTIFIER}(?:\s*\.\s*(?:{IDENTIFIER}|(?=\s*\.)))*'

    def __init__(self, input_excel: str, output_excel: str, sheet_name: Optional[str] = None):
        if not load_workbook or not Workbook:
            raise RuntimeError("openpyxl non installato. Installa con 'pip install openpyxl'.")
        if not input_excel or not os.path.exists(input_excel):
            raise FileNotFoundError(f"File Excel di input non trovato: {input_excel}")
        
        self.input_excel = input_excel
        self.output_excel = output_excel
        self.sheet_name = sheet_name

    @staticmethod
    def _strip_delimiters(name: str) -> str:
        """Rimuove delimitatori da un identifier SQL."""
        name = name.strip()
        if name.startswith('[') and name.endswith(']'):
            return name[1:-1]
        elif name.startswith('"') and name.endswith('"'):
            return name[1:-1]
        elif name.startswith('`') and name.endswith('`'):
            return name[1:-1]
        return name

    @staticmethod
    def _normalize_identifier(name: str) -> str:
        """Normalizza un identifier per il confronto (lowercase, senza delimitatori)."""
        return SQLClauseAnalyzer._strip_delimiters(name).lower()

    @staticmethod
    def _strip_sql_comments(sql: str) -> str:
        """Rimuove commenti SQL (-- e /* */)."""
        # Rimuovi commenti multilinea /* */
        sql = re.sub(r'/\*.*?\*/', ' ', sql, flags=re.DOTALL)
        # Rimuovi commenti singola linea --
        sql = re.sub(r'--[^\n]*', ' ', sql)
        return sql

    def _extract_table_name_parts(self, qualified_name: str) -> Tuple[str, ...]:
        """
        Estrae le parti di un nome qualificato e le normalizza.
        Ritorna tupla con le parti normalizzate (da 1 a 4 elementi).
        """
        parts = [self._normalize_identifier(p.strip()) 
                 for p in qualified_name.split('.') if p.strip() and p.strip() != '']
        return tuple(parts)

    def _matches_table(self, qualified_name: str, target_schema: str, target_table: str) -> bool:
        """
        Verifica se un nome qualificato corrisponde alla tabella target.
        Confronta schema.table o solo table.
        """
        parts = self._extract_table_name_parts(qualified_name)
        target_schema_norm = self._normalize_identifier(target_schema)
        target_table_norm = self._normalize_identifier(target_table)
        
        if not parts:
            return False
        
        # Caso 1: solo table
        if len(parts) == 1:
            return parts[0] == target_table_norm
        
        # Caso 2: schema.table
        if len(parts) == 2:
            return parts[0] == target_schema_norm and parts[1] == target_table_norm
        
        # Caso 3: db.schema.table o server.db.schema.table
        # Prendiamo gli ultimi 2 elementi (schema, table)
        if len(parts) >= 3:
            return parts[-2] == target_schema_norm and parts[-1] == target_table_norm
        
        return False

    def _find_sql_clauses(self, script: str, schema: str, table: str) -> List[str]:
        """
        Cerca tutte le occorrenze della tabella nello script e identifica
        la SQL clause utilizzata (FROM, JOIN, UPDATE, DELETE, INSERT INTO, etc.).
        Ritorna lista di clause trovate.
        """
        if not script:
            return []
        
        # Rimuovi commenti
        clean_script = self._strip_sql_comments(script)
        
        clauses_found = []
        
        # Pattern per diverse SQL clause
        # Ordine importante: pattern più specifici prima
        patterns = [
            # INSERT INTO
            (r'\bINSERT\s+INTO\s+(' + self.QUALIFIED_TABLE + r')\b', 'INSERT INTO'),
            
            # DELETE FROM
            (r'\bDELETE\s+(?:FROM\s+)?(' + self.QUALIFIED_TABLE + r')\b', 'DELETE FROM'),
            
            # UPDATE
            (r'\bUPDATE\s+(' + self.QUALIFIED_TABLE + r')\b', 'UPDATE'),
            
            # MERGE INTO
            (r'\bMERGE\s+(?:INTO\s+)?(' + self.QUALIFIED_TABLE + r')\b', 'MERGE INTO'),
            
            # TRUNCATE TABLE
            (r'\bTRUNCATE\s+TABLE\s+(' + self.QUALIFIED_TABLE + r')\b', 'TRUNCATE TABLE'),
            
            # SELECT INTO
            (r'\bSELECT\s+.+?\s+INTO\s+(' + self.QUALIFIED_TABLE + r')\b', 'SELECT INTO'),
            
            # ALTER TABLE
            (r'\bALTER\s+TABLE\s+(' + self.QUALIFIED_TABLE + r')\b', 'ALTER TABLE'),
            
            # CREATE TABLE
            (r'\bCREATE\s+TABLE\s+(' + self.QUALIFIED_TABLE + r')\b', 'CREATE TABLE'),
            
            # DROP TABLE
            (r'\bDROP\s+TABLE\s+(?:IF\s+EXISTS\s+)?(' + self.QUALIFIED_TABLE + r')\b', 'DROP TABLE'),
            
            # Vari tipi di JOIN
            (r'\bFULL\s+OUTER\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'FULL OUTER JOIN'),
            (r'\bLEFT\s+OUTER\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'LEFT OUTER JOIN'),
            (r'\bRIGHT\s+OUTER\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'RIGHT OUTER JOIN'),
            (r'\bLEFT\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'LEFT JOIN'),
            (r'\bRIGHT\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'RIGHT JOIN'),
            (r'\bINNER\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'INNER JOIN'),
            (r'\bCROSS\s+JOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'CROSS JOIN'),
            (r'\bJOIN\s+(' + self.QUALIFIED_TABLE + r')\b', 'JOIN'),
            
            # CROSS APPLY / OUTER APPLY
            (r'\bCROSS\s+APPLY\s+(' + self.QUALIFIED_TABLE + r')\b', 'CROSS APPLY'),
            (r'\bOUTER\s+APPLY\s+(' + self.QUALIFIED_TABLE + r')\b', 'OUTER APPLY'),
            
            # FROM (deve essere dopo i JOIN per evitare falsi positivi)
            (r'\bFROM\s+(' + self.QUALIFIED_TABLE + r')\b', 'FROM'),
        ]
        
        for pattern, clause_name in patterns:
            matches = re.finditer(pattern, clean_script, re.IGNORECASE | re.DOTALL)
            for match in matches:
                table_ref = match.group(1)
                if self._matches_table(table_ref, schema, table):
                    clauses_found.append(clause_name)
        
        # Rimuovi duplicati mantenendo l'ordine
        seen = set()
        unique_clauses = []
        for clause in clauses_found:
            if clause not in seen:
                seen.add(clause)
                unique_clauses.append(clause)
        
        return unique_clauses

    def _read_input_excel(self) -> List[Tuple[str, str, str, str, str, str, str]]:
        """
        Legge il file Excel di input.
        Ritorna lista di tuple: (server, database, schema, table, nome_oggetto, tipo_oggetto, script_creazione)
        """
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        
        rows_data = []
        
        # Legge intestazioni (prima riga)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        # Trova indici delle colonne (case-insensitive)
        def find_col_index(header_name: str) -> Optional[int]:
            header_lower = header_name.lower()
            for idx, h in enumerate(headers):
                if h and h.lower().strip() == header_lower:
                    return idx
            return None
        
        server_idx = find_col_index("server")
        db_idx = find_col_index("database")
        schema_idx = find_col_index("schema")
        table_idx = find_col_index("table")
        obj_name_idx = find_col_index("nome oggetto")
        obj_type_idx = find_col_index("tipo oggetto")
        script_idx = find_col_index("script creazione")
        
        if any(idx is None for idx in [server_idx, db_idx, schema_idx, table_idx, 
                                        obj_name_idx, obj_type_idx, script_idx]):
            wb.close()
            raise ValueError("Colonne richieste non trovate. Attese: server, database, schema, "
                           "table, nome oggetto, tipo oggetto, script creazione")
        
        # Legge dati
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= max(server_idx, db_idx, schema_idx, table_idx,
                                          obj_name_idx, obj_type_idx, script_idx):
                continue
            
            server = row[server_idx]
            database = row[db_idx]
            schema = row[schema_idx]
            table = row[table_idx]
            obj_name = row[obj_name_idx]
            obj_type = row[obj_type_idx]
            script = row[script_idx]
            
            if not all([database, schema, table, obj_name]):
                print(f"Riga {idx}: dati incompleti, saltata")
                continue
            
            rows_data.append((
                str(server).strip() if server else "",
                str(database).strip(),
                str(schema).strip(),
                str(table).strip(),
                str(obj_name).strip(),
                str(obj_type).strip() if obj_type else "",
                str(script) if script else ""
            ))
        
        wb.close()
        return rows_data

    def _create_output_excel(self, data: List[Tuple[str, str, str, str, str, str, str, str]]):
        """
        Crea il file Excel di output, splittando su più file se necessario.
        data: lista di tuple (server, database, schema, table, nome_oggetto, tipo_oggetto, script_creazione, sql_clause)
        """
        headers = [
            "Server",
            "Database",
            "Schema",
            "Table",
            "Nome Oggetto",
            "Tipo Oggetto",
            "Script Creazione",
            "SQL Clause",
        ]
        widths = [15, 20, 15, 30, 40, 25, 100, 30]

        try:
            from Report.Excel_Writer import write_rows_split_across_files
        except Exception:
            write_rows_split_across_files = None  # type: ignore

        if write_rows_split_across_files is not None:
            written = write_rows_split_across_files(
                headers=headers,
                rows=data,
                base_output_path=self.output_excel,
                sheet_name="Analisi SQL Clause",
                column_widths=widths,
            )
            print(f"File di output creato: {', '.join(written)}")
        else:
            # Fallback a singolo file
            wb = Workbook()
            ws = wb.active
            ws.title = "Analisi SQL Clause"
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = cell.font.copy(bold=True)
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            from openpyxl.utils import get_column_letter
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            wb.save(self.output_excel)
            print(f"File di output creato: {self.output_excel}")

    def process(self):
        """Elabora il file Excel e genera l'output con le SQL clause."""
        print("Lettura file Excel di input...")
        input_data = self._read_input_excel()
        print(f"Trovate {len(input_data)} righe da analizzare")
        
        output_data = []
        
        for idx, (server, database, schema, table, obj_name, obj_type, script) in enumerate(input_data, start=1):
            print(f"[{idx}/{len(input_data)}] Analisi: {obj_name} per tabella {schema}.{table}")
            
            # Trova le SQL clause
            clauses = self._find_sql_clauses(script, schema, table)
            
            if clauses:
                # Crea una riga per ogni clause trovata (o una singola riga con tutte le clause)
                clause_str = ", ".join(clauses)
                output_data.append((server, database, schema, table, obj_name, 
                                  obj_type, script, clause_str))
                print(f"  Trovate clause: {clause_str}")
            else:
                # Nessuna clause trovata, ma mantieni la riga
                output_data.append((server, database, schema, table, obj_name, 
                                  obj_type, script, "Non trovata"))
                print(f"  Nessuna clause trovata")
        
        print(f"\nCreazione file di output...")
        self._create_output_excel(output_data)
        print(f"\nElaborazione completata! Totale righe elaborate: {len(output_data)}")


def main():
    """Funzione principale."""
    if not INPUT_EXCEL_PATH:
        print("ERRORE: INPUT_EXCEL_PATH non configurato!")
        print("Modifica lo script e imposta INPUT_EXCEL_PATH con il percorso del file Excel di input.")
        return
    
    if not OUTPUT_EXCEL_PATH:
        print("ERRORE: OUTPUT_EXCEL_PATH non configurato!")
        print("Modifica lo script e imposta OUTPUT_EXCEL_PATH con il percorso del file Excel di output.")
        return
    
    try:
        analyzer = SQLClauseAnalyzer(INPUT_EXCEL_PATH, OUTPUT_EXCEL_PATH, INPUT_SHEET_NAME)
        analyzer.process()
    except Exception as e:
        print(f"ERRORE: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
