# -----------------------------------------------------------------------------
# Scopo: estrae le informazioni di connessione da un file Excel e produce
# nuovi file Excel che per ogni tabella individuata trovano gli oggetti
# associati (stored procedure, trigger, function, inline table valued function,
# table valued function). Gli oggetti sono considerati associati se referenziano
# la tabella di origine tramite FROM, JOIN, INSERT INTO, DELETE, ALTER, UPDATE, etc.
# Crea un file Excel ogni 50 connessioni analizzate.
# -----------------------------------------------------------------------------

import os
import re
from typing import List, Optional, Dict, Tuple, Set

try:
    import pyodbc
except Exception:
    pyodbc = None  # type: ignore

try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore

# -----------------------------------------------------------------------------
# CONFIGURAZIONE: inserisci qui i percorsi dei file
# -----------------------------------------------------------------------------

# Percorso del file Excel sorgente con le connessioni
INPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\path\to\tabelle_sorgente.xlsx"

# Percorso e nome base del file di destinazione (verrà aggiunto _1, _2, etc.)
OUTPUT_BASE_PATH: Optional[str] = None  # es: r"C:\path\to\oggetti_associati"

# Nome del foglio nel file Excel sorgente (None = primo foglio)
INPUT_SHEET_NAME: Optional[str] = None

# Server di default (configurabile se diverso da EPCP3)
DEFAULT_SERVER: str = "EPCP3"

# Numero di connessioni per file Excel
CONNECTIONS_PER_FILE: int = 50

# -----------------------------------------------------------------------------
# CONFIGURAZIONE CONNESSIONE SQL SERVER
# -----------------------------------------------------------------------------

ODBC_DRIVERS: List[str] = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
    "SQL Server Native Client 11.0",
    "ODBC Driver 13 for SQL Server",
    "ODBC Driver 11 for SQL Server",
]

TRUSTED_CONNECTION: bool = True
SQL_USERNAME: Optional[str] = None
SQL_PASSWORD: Optional[str] = None
CONNECTION_TEST_TIMEOUT: int = 3
QUERY_TIMEOUT: int = 60
ODBC_ENCRYPT_OPTS: str = "Encrypt=no;TrustServerCertificate=yes;"


class SQLObjectExtractor:
    """Estrae oggetti SQL (SP, trigger, function) associati a tabelle specificate in un Excel."""

    # Tipi di oggetti da cercare
    OBJECT_TYPES = {
        'P': 'Stored Procedure',
        'TR': 'Trigger',
        'FN': 'Scalar Function',
        'IF': 'Inline Table-Valued Function',
        'TF': 'Table-Valued Function'
    }

    def __init__(self, input_excel: str, output_base: str, sheet_name: Optional[str] = None):
        if not load_workbook or not Workbook:
            raise RuntimeError("openpyxl non installato. Installa con 'pip install openpyxl'.")
        if not pyodbc:
            raise RuntimeError("pyodbc non installato. Installa con 'pip install pyodbc'.")
        if not input_excel or not os.path.exists(input_excel):
            raise FileNotFoundError(f"File Excel sorgente non trovato: {input_excel}")
        
        self.input_excel = input_excel
        self.output_base = output_base
        self.sheet_name = sheet_name
        self.driver = self._get_available_driver()
        self.connections_cache: Dict[str, any] = {}

    def _get_available_driver(self) -> str:
        """Trova il primo driver ODBC disponibile."""
        available = pyodbc.drivers()
        for driver in ODBC_DRIVERS:
            if driver in available:
                return driver
        raise RuntimeError(f"Nessun driver ODBC trovato. Disponibili: {available}")

    def _get_connection(self, server: str, database: str):
        """Ottiene una connessione al database, riutilizzando se già presente nella cache."""
        cache_key = f"{server}|{database}"
        
        if cache_key in self.connections_cache:
            return self.connections_cache[cache_key]
        
        conn_str_parts = [
            f"DRIVER={{{self.driver}}}",
            f"SERVER={server}",
            f"DATABASE={database}",
            ODBC_ENCRYPT_OPTS
        ]
        
        if TRUSTED_CONNECTION:
            conn_str_parts.append("Trusted_Connection=yes")
        else:
            if SQL_USERNAME and SQL_PASSWORD:
                conn_str_parts.append(f"UID={SQL_USERNAME}")
                conn_str_parts.append(f"PWD={SQL_PASSWORD}")
        
        conn_str = ";".join(conn_str_parts)
        
        try:
            conn = pyodbc.connect(conn_str, timeout=CONNECTION_TEST_TIMEOUT)
            conn.timeout = QUERY_TIMEOUT
            self.connections_cache[cache_key] = conn
            return conn
        except Exception as e:
            raise RuntimeError(f"Errore connessione a {server}.{database}: {e}")

    def _read_input_excel(self) -> List[Tuple[str, str, str, str]]:
        """Legge il file Excel sorgente e restituisce lista di (server, db, schema, table)."""
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        
        tables = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            
            server = row[0] if row[0] else DEFAULT_SERVER
            database = row[1]
            schema = row[2]
            table = row[3]
            
            if not all([database, schema, table]):
                print(f"Riga {idx}: dati incompleti, saltata")
                continue
            
            tables.append((str(server).strip(), str(database).strip(), 
                          str(schema).strip(), str(table).strip()))
        
        wb.close()
        return tables

    def _find_associated_objects(self, server: str, database: str, 
                                 schema: str, table: str) -> List[Tuple[str, str, str]]:
        """
        Trova tutti gli oggetti (SP, trigger, function) che referenziano la tabella specificata.
        Ritorna lista di tuple: (nome_oggetto, tipo_oggetto, definizione_oggetto)
        """
        conn = self._get_connection(server, database)
        cursor = conn.cursor()
        
        # Query per trovare oggetti che referenziano la tabella
        # Usa sys.sql_expression_dependencies per trovare le dipendenze
        query = """
        SELECT DISTINCT
            o.name AS object_name,
            o.type AS object_type,
            OBJECT_DEFINITION(o.object_id) AS object_definition
        FROM sys.sql_expression_dependencies d
        INNER JOIN sys.objects o ON d.referencing_id = o.object_id
        INNER JOIN sys.objects ref ON d.referenced_id = ref.object_id
        INNER JOIN sys.schemas s ON ref.schema_id = s.schema_id
        WHERE s.name = ?
          AND ref.name = ?
          AND o.type IN ('P', 'TR', 'FN', 'IF', 'TF')
          AND OBJECT_DEFINITION(o.object_id) IS NOT NULL
        ORDER BY o.type, o.name
        """
        
        results = []
        try:
            cursor.execute(query, (schema, table))
            rows = cursor.fetchall()
            
            for row in rows:
                obj_name = row.object_name
                obj_type_code = row.object_type.strip()
                obj_definition = row.object_definition
                
                obj_type_desc = self.OBJECT_TYPES.get(obj_type_code, f"Unknown ({obj_type_code})")
                
                if obj_definition:
                    results.append((obj_name, obj_type_desc, obj_definition))
        
        except Exception as e:
            print(f"Errore query per {server}.{database}.{schema}.{table}: {e}")
        
        finally:
            cursor.close()
        
        return results

    def _create_output_excel(self, data: List[Tuple[str, str, str, str]], file_num: int):
        """Crea un file Excel di output con i dati specificati."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Oggetti Associati"
        
        # Intestazioni
        ws['A1'] = "Connessione Origine"
        ws['B1'] = "Nome Oggetto"
        ws['C1'] = "Tipo Oggetto"
        ws['D1'] = "Script Creazione"
        
        # Formattazione intestazioni
        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = ws[cell].font.copy(bold=True)
        
        # Dati
        for idx, (origin, obj_name, obj_type, obj_script) in enumerate(data, start=2):
            ws[f'A{idx}'] = origin
            ws[f'B{idx}'] = obj_name
            ws[f'C{idx}'] = obj_type
            ws[f'D{idx}'] = obj_script
        
        # Adatta larghezza colonne
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 100
        
        # Salva file
        output_path = f"{self.output_base}_{file_num}.xlsx"
        wb.save(output_path)
        print(f"Creato file: {output_path}")

    def process(self):
        """Elabora tutte le tabelle e crea i file Excel di output."""
        print("Lettura file Excel sorgente...")
        tables = self._read_input_excel()
        print(f"Trovate {len(tables)} tabelle da analizzare")
        
        all_results = []
        file_counter = 1
        
        for idx, (server, database, schema, table) in enumerate(tables, start=1):
            origin_connection = f"{server}.{database}.{schema}.{table}"
            print(f"[{idx}/{len(tables)}] Analisi: {origin_connection}")
            
            try:
                associated_objects = self._find_associated_objects(server, database, schema, table)
                
                if associated_objects:
                    for obj_name, obj_type, obj_script in associated_objects:
                        all_results.append((origin_connection, obj_name, obj_type, obj_script))
                    print(f"  Trovati {len(associated_objects)} oggetti associati")
                else:
                    print(f"  Nessun oggetto associato trovato")
            
            except Exception as e:
                print(f"  Errore: {e}")
                continue
            
            # Crea file ogni CONNECTIONS_PER_FILE connessioni
            if idx % CONNECTIONS_PER_FILE == 0 and all_results:
                self._create_output_excel(all_results, file_counter)
                all_results = []
                file_counter += 1
        
        # Crea file finale con eventuali risultati rimanenti
        if all_results:
            self._create_output_excel(all_results, file_counter)
        
        # Chiudi tutte le connessioni
        for conn in self.connections_cache.values():
            try:
                conn.close()
            except:
                pass
        
        print(f"\nElaborazione completata! Creati {file_counter} file di output.")


def main():
    """Funzione principale."""
    if not INPUT_EXCEL_PATH:
        print("ERRORE: INPUT_EXCEL_PATH non configurato!")
        print("Modifica lo script e imposta INPUT_EXCEL_PATH con il percorso del file Excel sorgente.")
        return
    
    if not OUTPUT_BASE_PATH:
        print("ERRORE: OUTPUT_BASE_PATH non configurato!")
        print("Modifica lo script e imposta OUTPUT_BASE_PATH con il percorso base per i file di output.")
        return
    
    try:
        extractor = SQLObjectExtractor(INPUT_EXCEL_PATH, OUTPUT_BASE_PATH, INPUT_SHEET_NAME)
        extractor.process()
    except Exception as e:
        print(f"ERRORE: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
