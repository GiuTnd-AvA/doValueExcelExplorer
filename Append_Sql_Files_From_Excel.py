import os
from typing import List, Optional

# Third-party dependency used elsewhere in the workspace
# We rely on openpyxl since the project already uses it.
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# -----------------------------------------------------------------------------
# Configurazione: INSERISCI QUI il percorso al file Excel che contiene la lista
# dei percorsi dei file .sql. Il foglio atteso è "Lista file SQL" con le colonne
# "Percorsi" e "File". Se il foglio/colonne differiscono, il parser prova
# a dedurre automaticamente la prima colonna come percorso.
# -----------------------------------------------------------------------------
EXCEL_LIST_PATH: Optional[str] = None  # es: r"C:\\Users\\...\\Report_Connessioni_0-99.xlsx"

# Percorso di output per il file TXT aggregato. Se None, usa la cartella
# dell'Excel di input e crea "SQL_Append.txt".
OUTPUT_TXT_PATH: Optional[str] = None

# Se True, crea anche una copia .sql affiancata al .txt
CREATE_SQL_COPY: bool = True


class SqlFilesAppender:
    """Legge un Excel con la lista dei percorsi dei file .sql e
    crea un file TXT con l'append di tutti i contenuti, separati da
    una riga di commento nel formato: --<progressivo> <percorso_file.sql>

    Il file risultante può essere aperto come .sql grazie ai commenti `--`.
    """

    def __init__(self, excel_path: str, output_txt: Optional[str] = None, sheet_name: Optional[str] = "Lista file SQL"):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if not excel_path:
            raise ValueError("Percorso Excel non valorizzato.")
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel non trovato: {excel_path}")
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        if output_txt:
            self.output_txt = output_txt
        else:
            base_dir = os.path.dirname(excel_path) or os.getcwd()
            self.output_txt = os.path.join(base_dir, "SQL_Append.txt")

    def _read_paths_from_excel(self) -> List[str]:
        """Legge le righe dell'Excel e restituisce percorsi completi a file .sql.

        Convenzione supportata (consigliata):
        - Colonna A: percorso directory
        - Colonna B: nome file (es: script.sql)

        Compatibilità:
        - Se in A c'è già il percorso completo che termina con .sql, usa A.
        - Se esiste un'intestazione "Percorsi", viene usata per la colonna dei percorsi.
        - In assenza di B, usa solo A.
        """
        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
        try:
            ws = wb[self.sheet_name] if self.sheet_name and self.sheet_name in wb.sheetnames else wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []

            # Determina se la prima riga è intestazione
            first = rows[0] or ()
            headers = [str(h).strip() if h is not None else "" for h in first]
            looks_like_header = any(h.lower() in ("percorsi", "file", "percorso", "nomefile") for h in headers)

            # Indici di default: A=0 (dir o path), B=1 (filename)
            start_row = 1 if looks_like_header else 0
            idx_a = 0
            idx_b = 1

            # Se c'è una colonna chiamata "Percorsi" la usiamo come A
            if "Percorsi" in headers:
                idx_a = headers.index("Percorsi")
            # Se c'è una colonna chiamata "File" la usiamo come B
            if "File" in headers:
                idx_b = headers.index("File")

            paths: List[str] = []
            for r in rows[start_row:]:
                if not r:
                    continue
                # Estrai A e B se presenti
                a = str(r[idx_a]).strip() if idx_a < len(r) and r[idx_a] is not None else ""
                b = str(r[idx_b]).strip() if idx_b < len(r) and r[idx_b] is not None else ""

                # Composizione logica del percorso
                candidate: str = ""
                if a:
                    a_exp = os.path.expanduser(os.path.expandvars(a))
                else:
                    a_exp = ""
                if b:
                    b_exp = os.path.expanduser(os.path.expandvars(b))
                else:
                    b_exp = ""

                if a_exp.lower().endswith('.sql'):
                    # Colonna A ha già un percorso completo a .sql
                    candidate = a_exp
                elif a_exp and b_exp:
                    # A è directory, B è filename
                    if a_exp.endswith(('\\', '/')):
                        candidate = a_exp + b_exp
                    else:
                        candidate = os.path.join(a_exp, b_exp)
                elif a_exp:
                    candidate = a_exp
                elif b_exp:
                    candidate = b_exp
                else:
                    continue

                candidate = os.path.normpath(candidate)
                if candidate:
                    paths.append(candidate)

            return paths
        finally:
            wb.close()

    def _read_file_text(self, path: str) -> str:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            try:
                with open(path, "r", encoding="latin-1", errors="ignore") as f:
                    return f.read()
            except Exception:
                return ""

    def run(self) -> str:
        paths = self._read_paths_from_excel()
        if not paths:
            raise RuntimeError("Nessun percorso trovato nell'Excel.")
        # Normalizza percorsi (espande %VAR% e ~) e filtra solo .sql
        norm_paths: List[str] = []
        for p in paths:
            np = os.path.expanduser(os.path.expandvars(p))
            if np.lower().endswith(".sql"):
                norm_paths.append(np)
        if not norm_paths:
            raise RuntimeError("Nessun file .sql trovato nella lista.")

        # Scrivi output aggregato
        out_dir = os.path.dirname(self.output_txt)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        with open(self.output_txt, "w", encoding="utf-8", errors="ignore") as out:
            for idx, fp in enumerate(norm_paths, start=1):
                # Separatore richiesto: --<n> <percorso>
                out.write(f"--{idx} {fp}\n")
                if os.path.exists(fp):
                    out.write(self._read_file_text(fp))
                else:
                    out.write(f"-- ATTENZIONE: file non trovato: {fp}\n")
                # Garantisce una newline tra file
                out.write("\n")

        # Crea copia .sql se richiesto
        if CREATE_SQL_COPY:
            sql_copy = os.path.splitext(self.output_txt)[0] + ".sql"
            try:
                # Copia bytes per evitare ricompressioni di newline
                with open(self.output_txt, "rb") as src, open(sql_copy, "wb") as dst:
                    dst.write(src.read())
            except Exception:
                # Silenzioso: il .txt rimane comunque disponibile
                pass

        return self.output_txt


if __name__ == "__main__":
    # Esempio di esecuzione stand-alone.
    if not EXCEL_LIST_PATH:
        raise SystemExit("Imposta EXCEL_LIST_PATH all'inizio del file.")
    app = SqlFilesAppender(EXCEL_LIST_PATH, OUTPUT_TXT_PATH)
    out_path = app.run()
    print(f"File aggregato creato: {out_path}")
    if CREATE_SQL_COPY:
        print(f"Copia .sql creata: {os.path.splitext(out_path)[0] + '.sql'}")
