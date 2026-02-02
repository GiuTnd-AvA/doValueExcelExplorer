"""
Estrae per una lista di oggetti (tabelle o viste):
- Gli oggetti che SCRIVONO (INSERT/UPDATE/DELETE/MERGE) sulla tabella o sulla vista
- Nel caso di VISTE: anche le sorgenti (tabelle/viste) usate per costruire la vista

Input atteso (Excel): colonne case-insensitive
  - Server
  - DB (o Database)
  - Schema
  - Object (o Table o View o Name)

Output (Excel): due fogli
  - Writers: Server | DB | Schema | Object | ObjectType | WriterName | WriterType | DMLType
  - ViewSources: Server | DB | Schema | View | SourceSchema | SourceName | SourceType

Dipendenze Python:
  pip install pyodbc pandas openpyxl
"""

import os
from typing import Any, Dict, List, Optional, Tuple

try:
    import pyodbc  # type: ignore
except Exception:
    pyodbc = None  # type: ignore

try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None  # type: ignore


# ------------------------ Configurazione base ------------------------
INPUT_EXCEL_PATH: Optional[str] = None   # es: r"C:\\path\\lista_oggetti.xlsx"
OUTPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\writers_e_view_sources.xlsx"
ROWS_PER_FILE: int = 20  # scrivi un file ogni N righe analizzate
DEFAULT_SERVER: str = "EPCP3"
DEFAULT_SERVER: str = "EPCP3"
DEFAULT_DB: str = "master"

ODBC_DRIVERS: List[str] = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
    "SQL Server Native Client 11.0",
    "ODBC Driver 13 for SQL Server",
    "ODBC Driver 11 for SQL Server",
]

TRUSTED_CONNECTION: bool = True
SQL_USERNAME: Optional[str] = None
SQL_PASSWORD: Optional[str] = None
CONNECTION_TEST_TIMEOUT: int = 3
QUERY_TIMEOUT: int = 60
ODBC_ENCRYPT_OPTS: str = "Encrypt=no;TrustServerCertificate=yes;"


def _derive_part_path(base_path: str, part_index: int) -> str:
    """Restituisce un path con suffisso _partN prima dell'estensione.

    Per part_index == 1, restituisce base_path per mantenere intatto il primo nome.
    """
    root, ext = os.path.splitext(base_path)
    if not ext:
        ext = ".xlsx"
    if part_index == 1:
        return root + ext
    return f"{root}_part{part_index}{ext}"


class WritersAndViewSourcesExtractor:
    def __init__(self, input_excel: str, output_excel: str, rows_per_file: int = ROWS_PER_FILE):
        if pyodbc is None:
            raise RuntimeError("pyodbc non installato. Esegui: pip install pyodbc")
        if pd is None:
            raise RuntimeError("pandas non installato. Esegui: pip install pandas openpyxl")
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Esegui: pip install openpyxl")
        if not input_excel or not os.path.exists(input_excel):
            raise FileNotFoundError(f"File Excel di input non trovato: {input_excel}")
        self.input_excel = input_excel
        self.output_excel = output_excel or os.path.join(
            os.path.dirname(input_excel) or os.getcwd(), "Writers_And_View_Sources.xlsx"
        )
        self.rows_per_file = max(1, int(rows_per_file))

    # ------------------------ Lettura input ------------------------
    def _read_items(self) -> List[Tuple[str, str, str, str]]:
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not rows:
                return []

            headers = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            # Colonne richieste: Server, DB/Database, Schema, Object/Table/View/Name
            idx_server = headers.index("server") if "server" in headers else None
            idx_db = headers.index("db") if "db" in headers else (headers.index("database") if "database" in headers else None)
            idx_schema = headers.index("schema") if "schema" in headers else None
            name_keys = [k for k in ("object", "table", "view", "name") if k in headers]
            idx_name = headers.index(name_keys[0]) if name_keys else None

            if None in (idx_server, idx_db, idx_schema, idx_name):
                raise RuntimeError(
                    "Il foglio deve contenere le colonne: Server, DB (o Database), Schema, Object (o Table/View/Name)."
                )

            items: List[Tuple[str, str, str, str]] = []
            for r in rows[1:]:
                if not r:
                    continue
                server = str(r[idx_server]).strip() if r[idx_server] else DEFAULT_SERVER
                db = str(r[idx_db]).strip() if r[idx_db] else DEFAULT_DB
                schema = str(r[idx_schema]).strip() if r[idx_schema] else "dbo"
                name = str(r[idx_name]).strip() if r[idx_name] else ""
                if not name:
                    continue
                items.append((server, db, schema, name))
            return items
        finally:
            wb.close()

    # ------------------------ Connessione ODBC ------------------------
    def _candidate_drivers(self) -> List[str]:
        try:
            installed = [d for d in pyodbc.drivers() if "sql server" in d.lower()]
        except Exception:
            installed = []
        preferred = [
            "ODBC Driver 18 for SQL Server",
            "ODBC Driver 17 for SQL Server",
            "SQL Server",
            "SQL Server Native Client 11.0",
            "ODBC Driver 13 for SQL Server",
            "ODBC Driver 11 for SQL Server",
        ]
        ordered = [d for d in preferred if d in installed]
        ordered += [d for d in installed if d not in ordered]
        if not ordered:
            ordered = ODBC_DRIVERS
        print(f"[ODBC] Driver installati: {installed}")
        return ordered

    def _build_conn_str(self, server: str, db: str) -> str:
        last_error: Optional[Exception] = None
        tried: List[str] = []
        for drv in self._candidate_drivers():
            tried.append(drv)
            try:
                test_db = "master"
                enc_opts = ODBC_ENCRYPT_OPTS
                if drv.lower().strip() == "sql server":
                    enc_opts = ""
                test_conn_str = f"DRIVER={{{drv}}};SERVER={server};DATABASE={test_db};"
                if TRUSTED_CONNECTION:
                    test_conn_str += "Trusted_Connection=yes;"
                else:
                    if not SQL_USERNAME or not SQL_PASSWORD:
                        raise RuntimeError("Credenziali non impostate e Trusted_Connection disattivata.")
                    test_conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                test_conn_str += enc_opts
                conn = pyodbc.connect(test_conn_str, timeout=CONNECTION_TEST_TIMEOUT)
                conn.close()
                final_conn_str = f"DRIVER={{{drv}}};SERVER={server};DATABASE={db};"
                if TRUSTED_CONNECTION:
                    final_conn_str += "Trusted_Connection=yes;"
                else:
                    final_conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                final_conn_str += enc_opts
                print(f"[ODBC] Uso driver: {drv}")
                return final_conn_str
            except Exception as e:
                last_error = e
                continue
        available = ", ".join(pyodbc.drivers()) if pyodbc is not None else "pyodbc non disponibile"
        raise RuntimeError(
            f"Impossibile connettersi a {server}. Provati: {tried}. Installati: {available}. Ultimo errore: {last_error}"
        )

    # ------------------------ Info oggetto target ------------------------
    def _get_object_type(self, conn, schema: str, name: str) -> Tuple[str, str]:
        sql = (
            """
            SELECT type, type_desc
            FROM sys.objects
            WHERE object_id = OBJECT_ID(QUOTENAME(?) + N'.' + QUOTENAME(?));
            """
        )
        cur = conn.cursor()
        try:
            cur.execute(sql, (schema, name))
            r = cur.fetchone()
            if not r:
                return ("", "NOT_FOUND")
            code = str(r[0]) if r[0] is not None else ""
            desc = str(r[1]) if r[1] is not None else ""
            return (code, desc)
        except Exception:
            return ("", "ERROR")

    # ------------------------ Writers (chi scrive) ------------------------
    def _find_writers(self, conn, schema: str, name: str) -> List[Tuple[str, str, str]]:
        """Ritorna lista di (writer_name, writer_type_desc, dml_type) per oggetti che scrivono sull'oggetto target.
        Cerca nei moduli (SP/trigger/funzioni) i token DML: INSERT INTO / UPDATE / DELETE FROM / MERGE INTO.
        """
        # Cerco oggetti che referenziano il target, e filtro sul testo della definizione per DML
        # Nota: sys.sql_expression_dependencies trova i referencing_id. Filtriamo poi su definition sm.definition
        sql = (
            """
            SELECT DISTINCT o.name, o.type_desc, sm.definition
            FROM sys.sql_expression_dependencies d
            JOIN sys.objects o ON d.referencing_id = o.object_id
            JOIN sys.sql_modules sm ON sm.object_id = o.object_id
            WHERE d.referenced_id = OBJECT_ID(QUOTENAME(?) + N'.' + QUOTENAME(?))
              AND o.type IN ('P','TR','FN','IF','TF')
            """
        )
        cur = conn.cursor()
        writers: List[Tuple[str, str, str]] = []
        try:
            cur.execute(sql, (schema, name))
            rows = cur.fetchall()
            # Costruiamo pattern ragionevoli (senza usare LIKE qui, filtriamo in Python per coprire casi vari)
            target_variants = [
                f"[{schema}].[{name}]",
                f"{schema}.{name}",
                f"[{schema}].{name}",
                f"{schema}.[{name}]",
                f"{name}",  # fallback (può produrre falsi positivi)
            ]
            for r in rows:
                oname = str(r[0])
                otype = str(r[1])
                definition = str(r[2]) if r[2] is not None else ""
                text_lower = definition.lower()

                # Individua DML specifico rivolto al target
                dml_found: List[str] = []
                for variant in target_variants:
                    v = variant.lower()
                    if f"insert into {v}" in text_lower:
                        dml_found.append("INSERT")
                    if f"update {v}" in text_lower:
                        dml_found.append("UPDATE")
                    if f"delete from {v}" in text_lower:
                        dml_found.append("DELETE")
                    if f"merge into {v}" in text_lower or f"merge {v}" in text_lower:
                        dml_found.append("MERGE")

                for dml in sorted(set(dml_found)):
                    writers.append((oname, otype, dml))
        except Exception:
            return writers
        finally:
            try:
                cur.close()
            except Exception:
                pass
        return writers

    # ------------------------ Sorgenti di una vista ------------------------
    def _find_view_sources(self, conn, schema: str, view: str) -> List[Tuple[str, str, str]]:
        """Ritorna lista di (source_schema, source_name, source_type_desc) per le dipendenze di una vista."""
        sql = (
            """
            SELECT DISTINCT s2.name AS source_schema, ref.name AS source_name, ref.type_desc AS source_type_desc
            FROM sys.sql_expression_dependencies d
            JOIN sys.objects ref ON d.referenced_id = ref.object_id
            JOIN sys.schemas s2 ON ref.schema_id = s2.schema_id
            WHERE d.referencing_id = OBJECT_ID(QUOTENAME(?) + N'.' + QUOTENAME(?))
              AND ref.type IN ('U','V')
            """
        )
        cur = conn.cursor()
        sources: List[Tuple[str, str, str]] = []
        try:
            cur.execute(sql, (schema, view))
            for r in cur.fetchall():
                sources.append((str(r[0]), str(r[1]), str(r[2])))
        except Exception:
            return sources
        finally:
            try:
                cur.close()
            except Exception:
                pass
        return sources

    # ------------------------ Esecuzione ------------------------
    def run(self) -> str:
        items = self._read_items()
        if not items:
            raise RuntimeError("Nessun oggetto valido trovato nell'Excel di input.")

        total = len(items)
        print(f"[DEP] Totale oggetti da elaborare: {total}")
        writers_rows: List[List[str]] = []
        view_src_rows: List[List[str]] = []
        part: int = 1
        conns: Dict[Tuple[str, str], Any] = {}
        try:
            for idx, (server, db, schema, name) in enumerate(items, start=1):
                print(f"[DEP] {idx}/{total}: {server}.{db}.{schema}.{name}")
                key = (server, db)
                if key not in conns:
                    try:
                        conn_str = self._build_conn_str(server, db)
                        conns[key] = pyodbc.connect(conn_str, timeout=QUERY_TIMEOUT)
                    except Exception as e:
                        # registra errori di connessione in writers con tipo Connessione fallita
                        writers_rows.append([server, db, schema, name, "Sconosciuto", "Connessione fallita", "", f"ERROR: {e}"])
                        continue

                # Tipo oggetto target
                code, desc = self._get_object_type(conns[key], schema, name)
                obj_type = desc or code or "Sconosciuto"

                # Trova writers
                for wname, wtype, dml in self._find_writers(conns[key], schema, name):
                    writers_rows.append([server, db, schema, name, obj_type, wname, wtype, dml])

                # Se è vista, troviamo anche le sorgenti
                if code.upper() == "V" or obj_type.upper().startswith("VIEW"):
                    for sschema, sname, stype in self._find_view_sources(conns[key], schema, name):
                        view_src_rows.append([server, db, schema, name, sschema, sname, stype])
                # Scrivi chunk ogni N righe analizzate
                if idx % self.rows_per_file == 0:
                    out_path = _derive_part_path(self.output_excel, part)
                    self._write_chunk(out_path, writers_rows, view_src_rows)
                    print(f"[DEP] Creato file: {out_path}")
                    writers_rows.clear()
                    view_src_rows.clear()
                    part += 1
        finally:
            for c in conns.values():
                try:
                    c.close()
                except Exception:
                    pass

        # Scrivi l'ultimo chunk (se presente)
        if writers_rows or view_src_rows:
            out_path = _derive_part_path(self.output_excel, part)
            self._write_chunk(out_path, writers_rows, view_src_rows)
            print(f"[DEP] Creato file: {out_path}")

        return self.output_excel

    def _write_chunk(self, out_path: str, writers_rows: List[List[str]], view_src_rows: List[List[str]]) -> None:
        """Scrive un chunk in un file Excel con due fogli: Writers e ViewSources."""
        out_dir = os.path.dirname(out_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        writers_df = pd.DataFrame(
            writers_rows,
            columns=["Server", "DB", "Schema", "Object", "ObjectType", "WriterName", "WriterType", "DMLType"],
        )
        view_src_df = pd.DataFrame(
            view_src_rows,
            columns=["Server", "DB", "Schema", "View", "SourceSchema", "SourceName", "SourceType"],
        )

        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as w:
            if not writers_df.empty:
                writers_df.to_excel(w, index=False, sheet_name="Writers")
            else:
                pd.DataFrame(columns=["Server","DB","Schema","Object","ObjectType","WriterName","WriterType","DMLType"]).to_excel(w, index=False, sheet_name="Writers")
            if not view_src_df.empty:
                view_src_df.to_excel(w, index=False, sheet_name="ViewSources")
            else:
                pd.DataFrame(columns=["Server","DB","Schema","View","SourceSchema","SourceName","SourceType"]).to_excel(w, index=False, sheet_name="ViewSources")


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="Estrai writers e sorgenti di viste da una lista di oggetti.")
    parser.add_argument("--input", dest="input_excel", default=INPUT_EXCEL_PATH, help="Path Excel input (Server, DB, Schema, Object)")
    parser.add_argument("--output", dest="output_excel", default=OUTPUT_EXCEL_PATH, help="Path Excel di output base")
    parser.add_argument("--rows-per-file", dest="rows_per_file", type=int, default=ROWS_PER_FILE, help="Numero di righe analizzate per file generato")
    args = parser.parse_args()

    if not args.input_excel:
        print("ERRORE: specifica --input oppure configura INPUT_EXCEL_PATH in testa al file.")
        return
    out_path = args.output_excel or ""
    try:
        extractor = WritersAndViewSourcesExtractor(args.input_excel, out_path, rows_per_file=args.rows_per_file)
        out = extractor.run()
        print(f"Output scritto con base path: {out}")
    except Exception as e:
        print(f"ERRORE: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
