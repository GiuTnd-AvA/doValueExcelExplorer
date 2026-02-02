# -----------------------------------------------------------------------------
# Scopo: legge un file Excel con colonne:
# Nome Oggetto, Tipo Oggetto, Script Creazione
# Appende tutte le DDL in un unico file .sql, senza filtri sul tipo
# di oggetto. Ogni voce è preceduta da un commento con indice e nome oggetto.
# -----------------------------------------------------------------------------

import os
import re
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


# Config di default (puoi sovrascrivere da CLI se necessario)
DEFAULT_EXCEL_PATH: Optional[str] = None
DEFAULT_SHEET_NAME: Optional[str] = None  # usa il primo foglio se None
DEFAULT_OUTPUT_PATH: Optional[str] = None  # se None, crea output vicino all'Excel


class SPDDLAppender:
    """Legge un Excel e appende le DDL in un unico file .sql (nessun filtro).

    Richiede intestazioni (case-insensitive, spazi tollerati):
    - nome oggetto (alias: object name, nome, oggetto, nomeoggetto, objectname)
    - tipo oggetto (alias: object type, tipooggetto, objecttype)
    - script creazione (alias: ddl, scriptcreazione)
    """

    REQUIRED_HEADERS = [
        "nome oggetto",
        "tipo oggetto",
        "script creazione",
    ]
    HEADER_ALIASES = {
        "tipo oggetto": ["object type", "tipooggetto", "objecttype"],
        "script creazione": ["ddl", "scriptcreazione"],
        "nome oggetto": ["nome", "object name", "oggetto", "nomeoggetto", "objectname"],
    }

    def __init__(
        self,
        excel_path: str,
        output_txt: Optional[str] = None,
        sheet_name: Optional[str] = DEFAULT_SHEET_NAME,
    ):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if not excel_path:
            raise ValueError("Percorso Excel non valorizzato.")
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel non trovato: {excel_path}")
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        if output_txt:
            self.output_txt = output_txt
        else:
            base_dir = os.path.dirname(excel_path) or os.getcwd()
            self.output_txt = os.path.join(base_dir, "Objects_Append.sql")

    @staticmethod
    def _norm_header(h: Optional[str]) -> str:
        return (str(h).strip().lower() if h is not None else "")

    @staticmethod
    def _extract_name_from_ddl(ddl: str) -> Optional[str]:
        """Prova ad estrarre un nome oggetto da una DDL generica.

        Supporta pattern CREATE [OR ALTER] <tipo> [schema.]nome con o senza []
        per tipi comuni: view|procedure|function|table|trigger|synonym.
        """
        if not ddl:
            return None
        text = str(ddl)
        # Con brackets: [schema].[name] oppure [name]
        m = re.search(
            r"\bcreate\s+(?:or\s+alter\s+)?(?:view|procedure|function|table|trigger|synonym)\s+\[([^\]]+)\]\s*\.\s*\[([^\]]+)\]",
            text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if m:
            return m.group(2)
        m = re.search(
            r"\bcreate\s+(?:or\s+alter\s+)?(?:view|procedure|function|table|trigger|synonym)\s+\[([^\]]+)\]",
            text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if m:
            return m.group(1)
        # Senza brackets: schema.name oppure name
        m = re.search(
            r"\bcreate\s+(?:or\s+alter\s+)?(?:view|procedure|function|table|trigger|synonym)\s+([a-zA-Z0-9_]+)\s*\.\s*([a-zA-Z0-9_]+)\b",
            text,
            flags=re.IGNORECASE,
        )
        if m:
            return m.group(2)
        m = re.search(
            r"\bcreate\s+(?:or\s+alter\s+)?(?:view|procedure|function|table|trigger|synonym)\s+([a-zA-Z0-9_]+)\b",
            text,
            flags=re.IGNORECASE,
        )
        if m:
            return m.group(1)
        return None

    def _read_rows(self) -> List[Dict[str, str]]:
        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
        try:
            ws = (
                wb[self.sheet_name]
                if self.sheet_name and self.sheet_name in wb.sheetnames
                else wb.worksheets[0]
            )
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [self._norm_header(h) for h in (rows[0] or [])]
            idx_map: Dict[str, int] = {}
            for needed in self.REQUIRED_HEADERS:
                if needed in headers:
                    idx_map[needed] = headers.index(needed)
                    continue
                needed_compact = needed.replace(" ", "")
                found_idx = None
                for i, h in enumerate(headers):
                    if h.replace(" ", "") == needed_compact:
                        found_idx = i
                        break
                if found_idx is None:
                    for alias in self.HEADER_ALIASES.get(needed, []):
                        alias_norm = self._norm_header(alias)
                        alias_compact = alias_norm.replace(" ", "")
                        for i, h in enumerate(headers):
                            if h == alias_norm or h.replace(" ", "") == alias_compact:
                                found_idx = i
                                break
                        if found_idx is not None:
                            break
                if found_idx is not None:
                    idx_map[needed] = found_idx
                else:
                    raise RuntimeError(f"Colonna richiesta non trovata: '{needed}'")

            data_rows = rows[1:]
            out: List[Dict[str, str]] = []
            for r in data_rows:
                if not r:
                    continue
                def get(col: str) -> str:
                    i = idx_map[col]
                    return str(r[i]).strip() if i < len(r) and r[i] is not None else ""

                nome_oggetto = get("nome oggetto")
                tipo = get("tipo oggetto")
                ddl = get("script creazione")

                if (not nome_oggetto) and ddl:
                    extracted = self._extract_name_from_ddl(ddl)
                    if extracted:
                        nome_oggetto = extracted

                if not nome_oggetto:
                    nome_oggetto = "UNKNOWN_OBJECT"

                out.append(
                    {
                        "object_name": nome_oggetto,
                        "object_type": tipo,
                        "ddl": ddl,
                    }
                )
            return out
        finally:
            wb.close()

    def run(self) -> str:
        rows = self._read_rows()
        if not rows:
            raise RuntimeError("Nessun oggetto trovato nell'Excel.")

        # Forza estensione .sql se non presente
        out_path = self.output_txt
        root, ext = os.path.splitext(out_path)
        if ext.lower() != ".sql":
            out_path = root + ".sql"

        out_dir = os.path.dirname(out_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        with open(out_path, "w", encoding="utf-8", errors="ignore") as f:
            for idx, r in enumerate(rows, start=1):
                obj_name = r.get("object_name") or "UNKNOWN_OBJECT"
                obj_type = (r.get("object_type") or "").strip()
                suffix = f"\t{obj_type}.sql" if obj_type else ""
                header = f"--{idx} {obj_name}{suffix}"
                f.write(header + "\n")
                if r["ddl"]:
                    f.write(str(r["ddl"]))
                f.write("\n\n")

        return out_path


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description=(
            "Append DDL da Excel in un unico file .sql (nessun filtro). "
            "Intestazioni richieste: Nome Oggetto, Tipo Oggetto, Script Creazione."
        )
    )
    ap.add_argument("excel", nargs="?", help="Percorso al file Excel di input")
    ap.add_argument("-o", "--output", help="Percorso file di output .sql")
    ap.add_argument("-s", "--sheet", help="Nome foglio Excel da usare (default: primo)")
    args = ap.parse_args()

    excel_path = args.excel or DEFAULT_EXCEL_PATH
    if not excel_path:
        raise SystemExit(
            "Errore: specifica il percorso Excel (argomento 'excel') oppure imposta DEFAULT_EXCEL_PATH all'inizio del file."
        )

    app = SPDDLAppender(
        excel_path,
        args.output or DEFAULT_OUTPUT_PATH,
        args.sheet or DEFAULT_SHEET_NAME,
    )
    out = app.run()
    print(f"Output creato: {out}")
