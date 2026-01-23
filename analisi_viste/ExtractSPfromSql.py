# -------------------------------------------------------------------
# Scopo: prende un file .sql e lo legge per estrarre tutte le stored 
# procedures richiamate (EXEC, EXECUTE, sp_executesql) e create (CREATE/ALTER)
# -------------------------------------------------------------------

import argparse
import os
import re
import csv
import datetime
from typing import List, Dict, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

# User-configurable default input path
INPUT_SQL = r"c:/Users/giuseppe.tanda/Desktop/doValue/OneDrive_1_12-30-2025/Append SQL script.sql"

# Identifier pattern: [name] or "name" or `name` or simple names
IDENTIFIER = r'(?:\[[^\]]+\]|"[^"]+"|`[^`]+`|[A-Za-z_][A-Za-z0-9_$@]*)'
# Qualified names: sp_name or schema.sp_name or db.schema.sp_name
QUALIFIED = rf'{IDENTIFIER}(?:\s*\.\s*{IDENTIFIER}){{0,3}}'

# Detect patterns with missing schema to normalize to dbo
EMPTY_SCHEMA_3 = re.compile(rf'^(?P<db>{IDENTIFIER})\s*\.\s*\.\s*(?P<sp>{IDENTIFIER})$', re.IGNORECASE)
EMPTY_SCHEMA_4 = re.compile(rf'^(?P<server>{IDENTIFIER})\s*\.\s*(?P<db>{IDENTIFIER})\s*\.\s*\.\s*(?P<sp>{IDENTIFIER})$', re.IGNORECASE)

# CTE detection: to avoid matching CTEs as stored procedures
CTE_FIRST_PATTERN = re.compile(rf"\bWITH\s+(?P<name>{IDENTIFIER})\s*(?:\([^)]*\))?\s+AS\b", re.IGNORECASE|re.DOTALL)
CTE_NEXT_PATTERN = re.compile(rf",\s*(?P<name>{IDENTIFIER})\s*(?:\([^)]*\))?\s+AS\b", re.IGNORECASE|re.DOTALL)

def _strip_delimiters(s: str) -> str:
    """Remove brackets, quotes from identifier"""
    s = s.strip()
    if s.startswith('[') and s.endswith(']'):
        s = s[1:-1]
    elif s.startswith('"') and s.endswith('"'):
        s = s[1:-1]
    elif s.startswith('`') and s.endswith('`'):
        s = s[1:-1]
    return s

def _last_segment(qualified: str) -> str:
    """Get the last part of a qualified name"""
    parts = [p.strip() for p in qualified.split('.')]
    if not parts:
        return _strip_delimiters(qualified)
    return _strip_delimiters(parts[-1])

def _is_temp_sp(qualified: str) -> bool:
    """Check if it's a temp object (starts with #)"""
    last = _last_segment(qualified)
    return last.startswith('#') or last.startswith('@')

def _extract_cte_names(text: str) -> set:
    """Extract CTE names to avoid false positives"""
    names = set()
    for m in CTE_FIRST_PATTERN.finditer(text):
        names.add(_strip_delimiters(m.group('name')))
        tail = text[m.end():]
        for n in CTE_NEXT_PATTERN.finditer(tail):
            names.add(_strip_delimiters(n.group('name')))
    return {n.lower() for n in names}

def _strip_sql_comments(text: str) -> str:
    """Remove T-SQL style comments: line comments (--) and block comments (/* ... */)"""
    # Remove block comments
    text = re.sub(r"/\*[\s\S]*?\*/", " ", text)
    # Remove line comments
    text = re.sub(r"(?m)--.*$", "", text)
    return text

def _extract_sp_ddl(text: str, match_start: int, match_end: int) -> str:
    """Extract the complete DDL of a stored procedure starting from CREATE/ALTER PROCEDURE"""
    # Find the end of the procedure: either GO, next CREATE/ALTER/DROP, or end of text
    remaining = text[match_start:]
    
    # Pattern for finding the end: GO statement or next major DDL statement
    end_patterns = [
        r"(?i)^\s*GO\s*$",  # GO on its own line
        r"(?i)\bCREATE\s+(?:OR\s+ALTER\s+)?PROC(?:EDURE)?\b",
        r"(?i)\bALTER\s+PROC(?:EDURE)?\b",
        r"(?i)\bDROP\s+PROC(?:EDURE)?\b",
        r"(?i)\bCREATE\s+TABLE\b",
        r"(?i)\bCREATE\s+VIEW\b",
        r"(?i)\bCREATE\s+FUNCTION\b",
    ]
    
    # Skip the current match to avoid matching itself
    search_start = match_end - match_start
    
    # Find the earliest terminator
    min_end = len(remaining)
    for pattern in end_patterns:
        for m in re.finditer(pattern, remaining[search_start:], re.MULTILINE):
            end_pos = search_start + m.start()
            if end_pos > 10:  # Must be at least 10 chars after start to avoid matching self
                min_end = min(min_end, end_pos)
                break  # Only need first match for each pattern
    
    # Extract the DDL from match_start to the found end
    ddl = remaining[:min_end].strip()
    return ddl

# Patterns for stored procedure calls
CLAUSE_PATTERNS: List[Tuple[str, re.Pattern]] = [
    # CREATE PROCEDURE
    ("CREATE PROCEDURE", re.compile(rf"\bCREATE\s+(?:OR\s+ALTER\s+)?PROC(?:EDURE)?\s+(?P<sp>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    
    # ALTER PROCEDURE
    ("ALTER PROCEDURE", re.compile(rf"\bALTER\s+PROC(?:EDURE)?\s+(?P<sp>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    
    # DROP PROCEDURE
    ("DROP PROCEDURE", re.compile(rf"\bDROP\s+PROC(?:EDURE)?\s+(?:IF\s+EXISTS\s+)?(?P<sp>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    
    # EXEC/EXECUTE: cattura il nome della SP dopo EXEC
    ("EXEC", re.compile(rf"\bEXEC(?:UTE)?\s+(?!sp_executesql\b)(?P<sp>{QUALIFIED})", re.IGNORECASE|re.DOTALL)),
    
    # sp_executesql (parametrizzata)
    ("sp_executesql", re.compile(r"\bsp_executesql\b", re.IGNORECASE)),
]

MARKER_REGEX = re.compile(r"(?m)^\s*--\s*(?P<num>\d+)\s+(?P<path>.+?\.sql)\s*$")

# Detect variables in dynamic SQL
VARIABLE_PATTERN = re.compile(r"@[A-Za-z_][A-Za-z0-9_]*")


def extract_sp_matches(text: str, original_text: str = None) -> List[Dict[str, str]]:
    """Extract stored procedure references from SQL text
    
    Args:
        text: SQL text with comments stripped
        original_text: Original SQL text (with comments) for DDL extraction
    """
    if original_text is None:
        original_text = text
    
    collected: List[Dict[str, str]] = []
    cte_names = _extract_cte_names(text)
    
    for clause, pattern in CLAUSE_PATTERNS:
        for m in pattern.finditer(text):
            c = clause
            ddl = None
            
            # Special handling for sp_executesql (no sp name to extract)
            if clause == "sp_executesql":
                pos = m.start()
                collected.append({'Clause': c, 'StoredProcedure': 'sp_executesql', 'DDL': None, '_pos': pos})
                continue
            
            sp = m.group('sp').strip()
            
            # Extract DDL for CREATE/ALTER PROCEDURE
            if clause in ("CREATE PROCEDURE", "ALTER PROCEDURE"):
                try:
                    ddl = _extract_sp_ddl(original_text, m.start(), m.end())
                except Exception as e:
                    print(f"Warning: Could not extract DDL for {sp}: {e}")
                    ddl = None
            
            # Skip if part of dynamic SQL (variable concatenation)
            try:
                match_end = m.end('sp')
                following = text[match_end:match_end+20]
                following_stripped = following.lstrip()
                if (following_stripped.startswith("'") or 
                    following_stripped.startswith("+ @")):
                    continue
            except Exception:
                pass
            
            # Skip if contains variable (e.g., @sp_name)
            if VARIABLE_PATTERN.match(_strip_delimiters(sp)):
                continue
            
            # Normalize missing schema: db..sp or server.db..sp
            em3 = EMPTY_SCHEMA_3.match(sp)
            em4 = EMPTY_SCHEMA_4.match(sp)
            if em4:
                sp = f"{em4.group('server')}.{em4.group('db')}.dbo.{em4.group('sp')}"
            elif em3:
                sp = f"{em3.group('db')}.dbo.{em3.group('sp')}"
            
            # Skip temp objects and CTEs
            if _is_temp_sp(sp):
                continue
            if _strip_delimiters(_last_segment(sp)).lower() in cte_names:
                continue
            
            # Get position
            try:
                pos = m.start('sp')
            except Exception:
                pos = m.start()
            
            collected.append({'Clause': c, 'StoredProcedure': sp, 'DDL': ddl, '_pos': pos})
    
    # Sort by position to preserve encounter order
    collected.sort(key=lambda x: x['_pos'])
    
    # Drop position before returning
    return [{'Clause': x['Clause'], 'StoredProcedure': x['StoredProcedure'], 'DDL': x['DDL']} for x in collected]


def parse_blocks(content: str, input_path: str, verbose: bool = True) -> List[Dict[str, str]]:
    """Parse SQL content for stored procedure references, handling marker-separated blocks"""
    rows: List[Dict[str, str]] = []
    markers = list(MARKER_REGEX.finditer(content))
    
    if not markers:
        # Single block case
        if verbose:
            print("Nessun marker trovato; elaboro il file come singolo blocco…")
        file_name = os.path.basename(input_path)
        content_nc = _strip_sql_comments(content)
        matches = extract_sp_matches(content_nc, content)
        if verbose:
            print(f"  Stored procedures estratte: {len(matches)}")
        for m in matches:
            rows.append({
                'Path': input_path, 
                'File': file_name, 
                'Clause': m['Clause'], 
                'StoredProcedure': m['StoredProcedure'],
                'DDL': m['DDL']
            })
        return rows
    
    # Multiple blocks
    if verbose:
        print(f"Trovati {len(markers)} marker di file; elaboro i blocchi…")
    for i, marker in enumerate(markers):
        path = marker.group('path').strip()
        file_name = os.path.basename(path)
        start = marker.end()
        end = markers[i+1].start() if i < len(markers)-1 else len(content)
        block = content[start:end]
        block_nc = _strip_sql_comments(block)
        matches = extract_sp_matches(block_nc, block)
        if verbose:
            print(f"  Blocco {i+1}/{len(markers)}: {file_name}")
            print(f"    Stored procedures estratte: {len(matches)}")
        for m in matches:
            rows.append({
                'Path': path, 
                'File': file_name, 
                'Clause': m['Clause'], 
                'StoredProcedure': m['StoredProcedure'],
                'DDL': m['DDL']
            })
    
    return rows


def write_csv(rows: List[Dict[str, str]], output_path: str) -> None:
    """Write results to CSV file"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Path', 'File', 'Clause', 'StoredProcedure'])
        for r in rows:
            w.writerow([r['Path'], r['File'], r['Clause'], r['StoredProcedure']])


def write_xlsx(rows: List[Dict[str, str]], output_path: str) -> None:
    """Write results to Excel file"""
    if openpyxl is None:
        raise RuntimeError('openpyxl is not installed')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'StoredProcedures'

    # Headers
    ws.cell(row=1, column=1, value='Path')
    ws.cell(row=1, column=2, value='File')
    ws.cell(row=1, column=3, value='Clause')
    ws.cell(row=1, column=4, value='StoredProcedure')

    for idx, r in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=r['Path'])
        ws.cell(row=idx, column=2, value=r['File'])
        ws.cell(row=idx, column=3, value=r['Clause'])
        ws.cell(row=idx, column=4, value=r['StoredProcedure'])

    # Autofit columns
    for col in range(1, 5):
        max_len = 0
        for row in range(1, len(rows)+2):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 80)

    try:
        wb.save(output_path)
    except PermissionError:
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        base, ext = os.path.splitext(output_path)
        alt_path = f"{base}_{ts}{ext}"
        print(f"File Excel destinazione bloccato: salvo come {alt_path}")
        wb.save(alt_path)


def main():
    ap = argparse.ArgumentParser(description='Extract stored procedure references from SQL into Excel/CSV.')
    ap.add_argument('input', nargs='?', help='Path to SQL file (optional if INPUT_SQL is set)')
    ap.add_argument('-o', '--output', help='Output file path (.xlsx or .csv)')
    ap.add_argument('-f', '--format', choices=['xlsx', 'csv'], default='xlsx', help='Output format')
    args = ap.parse_args()

    # Resolve input path
    input_path = args.input if args.input else INPUT_SQL
    if not input_path:
        raise FileNotFoundError('Percorso input non fornito: passa l\'argomento "input" oppure imposta INPUT_SQL all\'inizio dello script.')
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'Input file not found: {input_path}')

    # Default output path
    if not args.output:
        ext = 'xlsx' if args.format == 'xlsx' else 'csv'
        args.output = os.path.join(os.path.dirname(input_path), f'SQL_StoredProcedure_References.{ext}')

    print('Avvio estrazione stored procedures…')
    print(f'Input SQL: {input_path}')
    with open(input_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    rows = parse_blocks(content, input_path, verbose=True)

    print(f'Totale stored procedures trovate: {len(rows)}')

    if args.format == 'csv':
        print(f'Scrittura CSV in {args.output}…')
        write_csv(rows, args.output)
        print(f'CSV written to {args.output}')
    else:
        print(f'Scrittura Excel in {args.output}…')
        write_xlsx(rows, args.output)
        print(f'Excel written to {args.output}')


if __name__ == '__main__':
    main()
