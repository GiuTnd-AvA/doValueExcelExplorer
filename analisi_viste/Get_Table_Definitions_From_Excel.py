# -----------------------------------------------------------------------------
# Scopo: legge un file con gli elementi di connessione e per ogni elementi mi dice
# il tipo di oggetto di DB che sta leggendo e nel database cerca la devinizione 
# -----------------------------------------------------------------------------

import os
from typing import List, Optional, Tuple, Dict, Any

try:
    import pyodbc
except Exception:
    pyodbc = None  # type: ignore

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore

try:
    import pandas as pd
except Exception:
    pd = None  # type: ignore

# -----------------------------------------------------------------------------
# Config: imposta i percorsi di input e output (Excel)
# -----------------------------------------------------------------------------
INPUT_EXCEL_PATH: Optional[str] = None   # es: r"C:\\path\\liste_tabelle.xlsx"
OUTPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\ddl_tabelle.xlsx"

# Default connessione
DEFAULT_SERVER: str = "EPCP3"
DEFAULT_DB: str = "master"

ODBC_DRIVERS: List[str] = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
    # legacy/common names
    "SQL Server Native Client 11.0",
    "ODBC Driver 13 for SQL Server",
    "ODBC Driver 11 for SQL Server",
]
TRUSTED_CONNECTION: bool = True
SQL_USERNAME: Optional[str] = None
SQL_PASSWORD: Optional[str] = None
CONNECTION_TEST_TIMEOUT: int = 3
QUERY_TIMEOUT: int = 60
# Regola se il server richiede Encrypt
ODBC_ENCRYPT_OPTS: str = "Encrypt=no;TrustServerCertificate=yes;"


class TableDefinitionExtractor:
    """Legge la lista tabelle da Excel (colonne: Server | DB | Schema | Table)
    e produce un Excel di output con le stesse colonne + DDL.
    """

    def __init__(self, input_excel: str, output_excel: str):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if pd is None:
            raise RuntimeError("pandas non installato. Installa 'pip install pandas openpyxl'.")
        if pyodbc is None:
            raise RuntimeError("pyodbc non installato. Installa 'pip install pyodbc'.")
        if not input_excel:
            raise ValueError("Imposta INPUT_EXCEL_PATH.")
        self.input_excel = input_excel
        self.output_excel = output_excel or os.path.join(os.path.dirname(input_excel) or os.getcwd(), "DDL_Tabelle.xlsx")

    # ---------------- Excel parsing ----------------
    def _read_items(self) -> List[Tuple[str, str, str, str]]:
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not rows:
                return []

            headers = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            # Serve almeno: server, db|database, schema, table
            required = {"server", "schema", "table"}
            has_db = ("db" in headers) or ("database" in headers)
            if not (required.issubset(set(headers)) and has_db):
                raise RuntimeError("Il foglio deve contenere le colonne: Server, DB (o Database), Schema, Table.")

            idx_server = headers.index("server")
            idx_db = headers.index("db") if "db" in headers else headers.index("database")
            idx_schema = headers.index("schema")
            idx_table = headers.index("table")

            items: List[Tuple[str, str, str, str]] = []
            for r in rows[1:]:
                if r is None:
                    continue
                server = str(r[idx_server]).strip() if r[idx_server] else DEFAULT_SERVER
                db = str(r[idx_db]).strip() if r[idx_db] else DEFAULT_DB
                schema = str(r[idx_schema]).strip() if r[idx_schema] else "dbo"
                table = str(r[idx_table]).strip() if r[idx_table] else ""
                if not table:
                    continue
                items.append((server, db, schema, table))
            return items
        finally:
            wb.close()

    # ---------------- Connection handling ----------------
    def _candidate_drivers(self) -> List[str]:
        try:
            installed = [d for d in pyodbc.drivers() if "sql server" in d.lower()]
        except Exception:
            installed = []
        preferred = [
            "ODBC Driver 18 for SQL Server",
            "ODBC Driver 17 for SQL Server",
            "SQL Server",
            "SQL Server Native Client 11.0",
            "ODBC Driver 13 for SQL Server",
            "ODBC Driver 11 for SQL Server",
        ]
        ordered = [d for d in preferred if d in installed]
        ordered += [d for d in installed if d not in ordered]
        if not ordered:
            ordered = ODBC_DRIVERS
        print(f"[ODBC] Driver installati: {installed}")
        return ordered

    def _build_conn_str(self, server: str, db: str) -> str:
        last_error: Optional[Exception] = None
        tried: List[str] = []
        for drv in self._candidate_drivers():
            tried.append(drv)
            try:
                # Per il test del driver usiamo sempre 'master' per evitare errori 4060 sul DB target
                test_db = "master"
                enc_opts = ODBC_ENCRYPT_OPTS
                # Alcuni driver legacy ("SQL Server") non supportano Encrypt/TrustServerCertificate
                if drv.lower().strip() == "sql server":
                    enc_opts = ""
                test_conn_str = f"DRIVER={{{drv}}};SERVER={server};DATABASE={test_db};"
                if TRUSTED_CONNECTION:
                    test_conn_str += "Trusted_Connection=yes;"
                else:
                    if not SQL_USERNAME or not SQL_PASSWORD:
                        raise RuntimeError("Credenziali non impostate e Trusted_Connection disattivata.")
                    test_conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                test_conn_str += enc_opts
                conn = pyodbc.connect(test_conn_str, timeout=CONNECTION_TEST_TIMEOUT)
                conn.close()
                # Driver valido: costruisco la stringa finale per il DB target
                final_conn_str = f"DRIVER={{{drv}}};SERVER={server};DATABASE={db};"
                if TRUSTED_CONNECTION:
                    final_conn_str += "Trusted_Connection=yes;"
                else:
                    final_conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                final_conn_str += enc_opts
                print(f"[ODBC] Uso driver: {drv}")
                return final_conn_str
            except Exception as e:
                last_error = e
                continue
        available = ", ".join(pyodbc.drivers()) if pyodbc is not None else "pyodbc non disponibile"
        raise RuntimeError(
            f"Impossibile stabilire connessione al server {server}. Provati driver: {tried}. Driver installati: {available}. Ultimo errore: {last_error}"
        )

    # ---------------- DDL builder ----------------
    def _fetch_table_ddl(self, conn, schema: str, table: str) -> str:
        tsql_stringagg = r"""
DECLARE @schema_table nvarchar(512) = QUOTENAME(?) + N'.' + QUOTENAME(?);
DECLARE @obj_id int = OBJECT_ID(@schema_table);
IF @obj_id IS NULL
    SELECT CAST(N'ERROR: tabella non trovata: ' + @schema_table AS nvarchar(max)) AS ddl;
ELSE
BEGIN
    DECLARE @cols nvarchar(max) =
    (
        SELECT STRING_AGG(
            N'[' + c.name + N'] ' +
            UPPER(t.name) +
            CASE 
                WHEN t.name IN (N'char',N'nchar',N'varchar',N'nvarchar',N'binary',N'varbinary') 
                     THEN N'(' + CASE 
                                   WHEN t.name IN (N'nchar',N'nvarchar') 
                                        THEN CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length/2 AS nvarchar(10)) END
                                   ELSE CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length AS nvarchar(10)) END
                                 END + N')'
                WHEN t.name IN (N'decimal',N'numeric') 
                     THEN N'(' + CAST(c.precision AS nvarchar(10)) + N',' + CAST(c.scale AS nvarchar(10)) + N')'
                WHEN t.name IN (N'datetime2',N'time',N'datetimeoffset')
                     THEN N'(' + CAST(c.scale AS nvarchar(10)) + N')'
                ELSE N''
            END +
            CASE WHEN ic.is_identity = 1 
                 THEN N' IDENTITY(' + CAST(ic.seed_value AS nvarchar(50)) + N',' + CAST(ic.increment_value AS nvarchar(50)) + N')' 
                 ELSE N'' 
            END +
            CASE WHEN c.is_nullable = 0 THEN N' NOT NULL' ELSE N' NULL' END +
            COALESCE(N' DEFAULT ' + dc.definition, N'')
            , N',' + CHAR(13) + CHAR(10)
        ) WITHIN GROUP (ORDER BY c.column_id)
        FROM sys.columns c
        JOIN sys.types t ON c.user_type_id = t.user_type_id
        LEFT JOIN sys.default_constraints dc ON dc.parent_object_id = c.object_id AND dc.parent_column_id = c.column_id
        LEFT JOIN sys.identity_columns ic ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        WHERE c.object_id = @obj_id
    );

    DECLARE @pk nvarchar(max) =
    (
        SELECT N'CONSTRAINT [' + kc.name + N'] PRIMARY KEY (' +
               STRING_AGG(N'[' + c.name + N']' + CASE WHEN ic.is_descending_key = 1 THEN N' DESC' ELSE N'' END, N', ')
               WITHIN GROUP (ORDER BY ic.key_ordinal) + N')'
        FROM sys.key_constraints kc
        JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
        JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id
        JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
        WHERE kc.type = 'PK' AND kc.parent_object_id = @obj_id
    );

    SELECT N'CREATE TABLE ' + @schema_table + CHAR(13) + CHAR(10) +
           N'(' + ISNULL(@cols, N'') + N')' +
           ISNULL(CHAR(13) + CHAR(10) + @pk, N'') AS ddl;
END
"""
        tsql_xmlpath = r"""
DECLARE @schema_table nvarchar(512) = QUOTENAME(?) + N'.' + QUOTENAME(?);
DECLARE @obj_id int = OBJECT_ID(@schema_table);
IF @obj_id IS NULL
    SELECT CAST(N'ERROR: tabella non trovata: ' + @schema_table AS nvarchar(max)) AS ddl;
ELSE
BEGIN
    DECLARE @cols nvarchar(max) = N'';
    SELECT @cols = STUFF((
        SELECT N',' + CHAR(13) + CHAR(10) +
               N'[' + c.name + N'] ' +
               UPPER(t.name) +
               CASE 
                   WHEN t.name IN (N'char',N'nchar',N'varchar',N'nvarchar',N'binary',N'varbinary') 
                        THEN N'(' + CASE 
                                      WHEN t.name IN (N'nchar',N'nvarchar') 
                                           THEN CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length/2 AS nvarchar(10)) END
                                      ELSE CASE WHEN c.max_length = -1 THEN N'MAX' ELSE CAST(c.max_length AS nvarchar(10)) END
                                    END + N')'
                   WHEN t.name IN (N'decimal',N'numeric') 
                        THEN N'(' + CAST(c.precision AS nvarchar(10)) + N',' + CAST(c.scale AS nvarchar(10)) + N')'
                   WHEN t.name IN (N'datetime2',N'time',N'datetimeoffset')
                        THEN N'(' + CAST(c.scale AS nvarchar(10)) + N')'
                   ELSE N''
               END +
               CASE WHEN ic.is_identity = 1 
                    THEN N' IDENTITY(' + CAST(ic.seed_value AS nvarchar(50)) + N',' + CAST(ic.increment_value AS nvarchar(50)) + N')' 
                    ELSE N'' 
               END +
               CASE WHEN c.is_nullable = 0 THEN N' NOT NULL' ELSE N' NULL' END +
               COALESCE(N' DEFAULT ' + dc.definition, N'')
        FROM sys.columns c
        JOIN sys.types t ON c.user_type_id = t.user_type_id
        LEFT JOIN sys.default_constraints dc ON dc.parent_object_id = c.object_id AND dc.parent_column_id = c.column_id
        LEFT JOIN sys.identity_columns ic ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        WHERE c.object_id = @obj_id
        ORDER BY c.column_id
        FOR XML PATH(''), TYPE
    ).value('.', 'nvarchar(max)'), 1, 1, N'');

    DECLARE @pk nvarchar(max) = NULL;
    SELECT @pk = N'CONSTRAINT [' + kc.name + N'] PRIMARY KEY (' +
                 STUFF((
                    SELECT N', ' + N'[' + c.name + N']' + CASE WHEN ic.is_descending_key = 1 THEN N' DESC' ELSE N'' END
                    FROM sys.index_columns ic
                    JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                    WHERE ic.object_id = i.object_id AND ic.index_id = i.index_id
                    ORDER BY ic.key_ordinal
                    FOR XML PATH(''), TYPE
                 ).value('.', 'nvarchar(max)'), 1, 2, N'') + N')'
    FROM sys.key_constraints kc
    JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
    WHERE kc.type = 'PK' AND kc.parent_object_id = @obj_id;

    SELECT N'CREATE TABLE ' + @schema_table + CHAR(13) + CHAR(10) +
           N'(' + ISNULL(@cols, N'') + N')' +
           ISNULL(CHAR(13) + CHAR(10) + @pk, N'') AS ddl;
END
"""
        cur = conn.cursor()
        try:
            cur.execute(tsql_stringagg, (schema, table))
        except Exception:
            cur.execute(tsql_xmlpath, (schema, table))
        row = cur.fetchone()
        return row[0] if row else ""

    def _get_object_type_info(self, conn, schema: str, name: str) -> Tuple[str, str, str]:
        """Ritorna (code, type_desc, label_per_output).
        Esempi: ('U','USER_TABLE','USER_TABLE'), ('V','VIEW','VIEW').
        """
        sql = (
            """
            SELECT type, type_desc
            FROM sys.objects
            WHERE object_id = OBJECT_ID(QUOTENAME(?) + '.' + QUOTENAME(?));
            """
        )
        cur = conn.cursor()
        try:
            cur.execute(sql, (schema, name))
            r = cur.fetchone()
            if not r:
                return ("", "NOT_FOUND", "Non trovato")
            code = str(r[0]) if r[0] is not None else ""
            desc = str(r[1]) if r[1] is not None else ""
            label = desc or code or "Sconosciuto"
            return (code, desc, label)
        except Exception:
            return ("", "ERROR", "Sconosciuto")

    def _fetch_view_definition(self, conn, schema: str, view: str) -> str:
        """Ritorna la definizione testuale della vista così come salvata nel DB.
        Usa sys.sql_modules/OBJECT_DEFINITION. Se la definizione non è disponibile
        (es. oggetto crittografato), restituisce un messaggio informativo.
        """
        sql = (
            """
            SELECT sm.definition
            FROM sys.sql_modules AS sm
            WHERE sm.object_id = OBJECT_ID(QUOTENAME(?) + N'.' + QUOTENAME(?));
            """
        )
        cur = conn.cursor()
        try:
            cur.execute(sql, (schema, view))
            r = cur.fetchone()
            if r and r[0]:
                return str(r[0])
            # fallback: OBJECT_DEFINITION
            cur.execute("SELECT OBJECT_DEFINITION(OBJECT_ID(QUOTENAME(?) + N'.' + QUOTENAME(?)))", (schema, view))
            r2 = cur.fetchone()
            if r2 and r2[0]:
                return str(r2[0])
            return "ERROR: definizione non disponibile (possibile oggetto crittografato)"
        except Exception as e:
            return f"ERROR: lettura definizione vista fallita: {e}"

    # ---------------- Main ----------------
    def run(self) -> str:
        items = self._read_items()
        if not items:
            raise RuntimeError("Nessuna tabella valida trovata nell'Excel di input.")

        total = len(items)
        print(f"[DDL] Totale tabelle da elaborare: {total}")
        results: List[List[str]] = []
        conns: Dict[Tuple[str, str], Any] = {}
        try:
            for idx, (server, db, schema, table) in enumerate(items, start=1):
                print(f"[DDL] Elaborazione {idx}/{total}: {server}.{db}.{schema}.{table}")
                key = (server, db)
                if key not in conns:
                    try:
                        conn_str = self._build_conn_str(server, db)
                        conns[key] = pyodbc.connect(conn_str, timeout=QUERY_TIMEOUT)
                    except Exception as e:
                        # Connessione non riuscita per questo server/db: si registra errore e si continua
                        results.append([server, db, schema, table, "ERROR", f"Connessione fallita: {e}"])
                        continue
                code, desc, obj_type = self._get_object_type_info(conns[key], schema, table)
                if code.upper() == "V" or desc.upper().startswith("VIEW") or obj_type.lower().startswith("vista"):
                    ddl = self._fetch_view_definition(conns[key], schema, table)
                else:
                    ddl = self._fetch_table_ddl(conns[key], schema, table)
                results.append([server, db, schema, table, obj_type, ddl])
        finally:
            for c in conns.values():
                try:
                    c.close()
                except Exception:
                    pass

        # Write output
        out_dir = os.path.dirname(self.output_excel)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        df = pd.DataFrame(results, columns=["Server", "DB", "Schema", "Table", "ObjectType", "DDL"])
        with pd.ExcelWriter(self.output_excel, engine="openpyxl", mode="w") as w:
            df.to_excel(w, index=False, sheet_name="DDL")
        return self.output_excel


if __name__ == "__main__":
    if not INPUT_EXCEL_PATH:
        raise SystemExit("Imposta INPUT_EXCEL_PATH a inizio file.")
    extractor = TableDefinitionExtractor(INPUT_EXCEL_PATH, OUTPUT_EXCEL_PATH or "")
    out = extractor.run()
    print(f"DDL scritto in: {out}")
