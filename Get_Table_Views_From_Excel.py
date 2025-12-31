import os
from typing import List, Optional, Tuple, Dict

try:
    import pyodbc
except Exception:
    pyodbc = None  # type: ignore

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore

try:
    import pandas as pd
except Exception:
    pd = None  # type: ignore

# -----------------------------------------------------------------------------
# Configurazione: inserisci qui i percorsi e i parametri di connessione.
# -----------------------------------------------------------------------------
INPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\Tabelle.xlsx"
OUTPUT_EXCEL_PATH: Optional[str] = None  # es: r"C:\\path\\VistePerTabella.xlsx"
DEFAULT_SERVER: str = "EPCP3"  # server costante richiesto
DEFAULT_DB: str = "master"
# Proveremo questi driver in ordine.
ODBC_DRIVERS: List[str] = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
    # legacy/common names
    "SQL Server Native Client 11.0",
    "ODBC Driver 13 for SQL Server",
    "ODBC Driver 11 for SQL Server",
]
TRUSTED_CONNECTION: bool = True
SQL_USERNAME: Optional[str] = None  # usato se TRUSTED_CONNECTION=False
SQL_PASSWORD: Optional[str] = None  # usato se TRUSTED_CONNECTION=False
# Timeout in secondi
QUERY_TIMEOUT: int = 60
# Timeout rapido per test connessione driver
CONNECTION_TEST_TIMEOUT: int = 3
# Salvataggio parziale ogni N elementi
PARTIAL_SAVE_EVERY: int = 50
# Opzioni di cifratura/Trust
ODBC_ENCRYPT_OPTS: str = "Encrypt=no;TrustServerCertificate=yes;"


class TableViewsExtractor:
    """Legge un Excel con colonne Server|DB|Schema|Table e per ogni tabella
    estrae tutte le viste che la referenziano con la loro definizione.

    Output: Excel con colonne [Server, DB, Schema, Table, Object_Name, Definition]
    """

    def __init__(self, input_excel: str, output_excel: str):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if pd is None:
            raise RuntimeError("pandas non installato. Installa 'pip install pandas openpyxl'.")
        if pyodbc is None:
            raise RuntimeError("pyodbc non installato. Installa 'pip install pyodbc'.")
        if not input_excel:
            raise ValueError("Percorso Excel di input non valorizzato.")
        self.input_excel = input_excel
        self.output_excel = output_excel or os.path.join(os.path.dirname(input_excel) or os.getcwd(), "VistePerTabella.xlsx")

    def _read_items(self) -> List[Tuple[str, str, str, str]]:
        """Ritorna lista di tuple (server, db, schema, table)."""
        wb = load_workbook(self.input_excel, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        items: List[Tuple[str, str, str, str]] = []
        if not rows:
            return items

        # Mappa header
        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        col_idx: Dict[str, int] = {}
        for name in ("server", "db", "database", "schema", "table"):
            if name in header:
                col_idx[name] = header.index(name)

        # Indici risolti
        idx_server = col_idx.get("server", -1)
        idx_db = col_idx.get("db", col_idx.get("database", -1))
        idx_schema = col_idx.get("schema", -1)
        idx_table = col_idx.get("table", -1)

        # Se non c'è header, prova layout semplice: Server|DB|Schema|Table
        start_row = 2 if idx_server != -1 or idx_db != -1 or idx_schema != -1 or idx_table != -1 else 1
        for r in rows[start_row - 1:]:
            try:
                server = str(r[idx_server]).strip() if idx_server != -1 and r[idx_server] is not None else DEFAULT_SERVER
                db = str(r[idx_db]).strip() if idx_db != -1 and r[idx_db] is not None else DEFAULT_DB
                schema = str(r[idx_schema]).strip() if idx_schema != -1 and r[idx_schema] is not None else "dbo"
                table = str(r[idx_table]).strip() if idx_table != -1 and r[idx_table] is not None else ""
            except Exception:
                # Fallback senza header: assumiamo ordine fisso
                vals = [v for v in r if v is not None]
                if len(vals) < 2:
                    continue
                server = DEFAULT_SERVER
                db = str(vals[0]).strip() if len(vals) >= 1 else DEFAULT_DB
                schema = str(vals[1]).strip() if len(vals) >= 2 else "dbo"
                table = str(vals[2]).strip() if len(vals) >= 3 else ""

            if not table:
                continue
            items.append((server or DEFAULT_SERVER, db or DEFAULT_DB, schema or "dbo", table))

        return items

    def _build_conn_str(self, server: str, db: str) -> str:
        """Replica la logica di connessione utilizzata in Get_Table_Definitions_From_Excel.
        Prova i driver in ODBC_DRIVERS in ordine, aggiunge Trusted/UID/PWD e le opzioni ODBC_ENCRYPT_OPTS,
        valida con una connessione rapida e ritorna la stringa valida.
        """
        last_error: Optional[Exception] = None
        for drv in ODBC_DRIVERS:
            try:
                conn_str = f"DRIVER={{{drv}}};SERVER={server};DATABASE={db};"
                if TRUSTED_CONNECTION:
                    conn_str += "Trusted_Connection=yes;"
                else:
                    if not SQL_USERNAME or not SQL_PASSWORD:
                        raise RuntimeError("Imposta SQL_USERNAME e SQL_PASSWORD oppure usa Trusted_Connection.")
                    conn_str += f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
                conn_str += ODBC_ENCRYPT_OPTS
                # Test rapido del driver
                tconn = pyodbc.connect(conn_str, timeout=CONNECTION_TEST_TIMEOUT)
                tconn.close()
                return conn_str
            except Exception as e:
                last_error = e
                continue
        raise RuntimeError(f"Nessun driver ODBC valido trovato. Ultimo errore: {last_error}")

    def _fetch_views_for_table(self, conn, schema: str, table: str) -> List[Tuple[str, str]]:
        """Ritorna lista di (view_name, definition) per la tabella schema.table.
        Usa sys.sql_expression_dependencies; se non trova nulla, fallback su ricerca nel testo della definizione.
        """
        # Prima: dipendenze espresse
        sql_dep = (
            """
            DECLARE @schema sysname = ?;
            DECLARE @table  sysname = ?;
            DECLARE @objId INT = OBJECT_ID(QUOTENAME(@schema) + '.' + QUOTENAME(@table));

            SELECT v.name AS view_name,
                   sm.definition
            FROM sys.sql_expression_dependencies AS d
            JOIN sys.objects AS v
                ON d.referencing_id = v.object_id AND v.type = 'V'
            JOIN sys.sql_modules AS sm
                ON v.object_id = sm.object_id
            WHERE d.referenced_id = @objId
            ORDER BY v.name;
            """
        )
        cur = conn.cursor()
        rows: List[Tuple[str, str]] = []
        try:
            cur.execute(sql_dep, (schema, table))
            for r in cur.fetchall():
                rows.append((str(r[0]), str(r[1])))
        except Exception:
            rows = []

        if rows:
            return rows

        # Fallback ROBUSTO: ricerca letterale con CHARINDEX per evitare falsi positivi dei bracket con LIKE
        sql_fb = (
            """
            DECLARE @schema sysname = ?;
            DECLARE @table  sysname = ?;

            DECLARE @two_br   nvarchar(400) = QUOTENAME(@schema) + N'.' + QUOTENAME(@table); -- [schema].[table]
            DECLARE @two_pl   nvarchar(400) = @schema + N'.' + @table;                         -- schema.table
            DECLARE @db       sysname       = DB_NAME();
            DECLARE @three_br nvarchar(600) = QUOTENAME(@db) + N'.' + @two_br;                 -- [db].[schema].[table]
            DECLARE @three_pl nvarchar(600) = @db + N'.' + @two_pl;                             -- db.schema.table

            SELECT DISTINCT v.name AS view_name,
                            sm.definition
            FROM sys.views AS v
            JOIN sys.sql_modules AS sm ON v.object_id = sm.object_id
            WHERE  CHARINDEX(@two_br ,  sm.definition) > 0
                OR CHARINDEX(@two_pl ,  sm.definition) > 0
                OR CHARINDEX(@three_br, sm.definition) > 0
                OR CHARINDEX(@three_pl, sm.definition) > 0
                OR CHARINDEX(N'FROM ' + @two_br ,  sm.definition) > 0
                OR CHARINDEX(N'JOIN ' + @two_br ,  sm.definition) > 0
                OR CHARINDEX(N'FROM ' + @two_pl ,  sm.definition) > 0
                OR CHARINDEX(N'JOIN ' + @two_pl ,  sm.definition) > 0
                OR CHARINDEX(N'FROM ' + @three_br, sm.definition) > 0
                OR CHARINDEX(N'JOIN ' + @three_br, sm.definition) > 0
                OR CHARINDEX(N'FROM ' + @three_pl, sm.definition) > 0
                OR CHARINDEX(N'JOIN ' + @three_pl, sm.definition) > 0
            ORDER BY v.name;
            """
        )
        cur = conn.cursor()
        try:
            cur.execute(sql_fb, (schema, table))
            fb_rows = cur.fetchall()
            rows = [(str(r[0]), str(r[1])) for r in fb_rows]
        except Exception:
            rows = []
        return rows

    def run(self) -> str:
        items = self._read_items()
        if not items:
            print("[VIEW] Nessuna tabella valida trovata nell'Excel di input.")
            return self.output_excel or ""

        total = len(items)
        print(f"[VIEW] Totale tabelle da processare: {total}")

        # Pool connessioni per server/db
        conns: Dict[Tuple[str, str], object] = {}
        results: List[List[str]] = []

        def _write_results(path: str, rows: List[List[str]]):
            out_dir = os.path.dirname(path)
            if out_dir and not os.path.exists(out_dir):
                os.makedirs(out_dir, exist_ok=True)
            df = pd.DataFrame(rows, columns=["Server", "DB", "Schema", "Table", "Object_Name", "Definition"])
            with pd.ExcelWriter(path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, index=False, sheet_name='Viste')

        try:
            for idx, (server, db, schema, table) in enumerate(items, start=1):
                print(f"[VIEW] Elaborazione {idx}/{total}: {server}.{db}.{schema}.{table}")

                key = (server, db)
                conn = conns.get(key)
                if conn is None:
                    conn_str = self._build_conn_str(server, db)
                    conn = pyodbc.connect(conn_str, timeout=QUERY_TIMEOUT)
                    conns[key] = conn

                views = self._fetch_views_for_table(conn, schema, table)
                if not views:
                    print(f"[VIEW] Nessuna vista trovata per {schema}.{table}")
                    continue
                print(f"[VIEW] Trovate {len(views)} viste per {schema}.{table}")
                for view_name, definition in views:
                    results.append([server, db, schema, table, view_name, definition])

                # Salvataggio parziale ogni PARTIAL_SAVE_EVERY elementi elaborati
                if idx % PARTIAL_SAVE_EVERY == 0:
                    base_dir = os.path.dirname(self.output_excel) or os.getcwd()
                    base_name = os.path.splitext(os.path.basename(self.output_excel or 'VistePerTabella.xlsx'))[0]
                    partial_path = os.path.join(base_dir, f"{base_name}_partial_{idx}.xlsx")
                    print(f"[VIEW] Salvataggio parziale {idx}/{total}: {partial_path}")
                    _write_results(partial_path, results)
        finally:
            # Chiudi connessioni
            for key, c in conns.items():
                try:
                    c.close()
                except Exception:
                    pass

        # Scrivi output
        if not results:
            print("[VIEW] Nessun risultato da scrivere.")
        _write_results(self.output_excel, results)
        print(f"[VIEW] Output scritto in: {self.output_excel}")
        return self.output_excel


if __name__ == "__main__":
    if not INPUT_EXCEL_PATH:
        raise SystemExit("Imposta INPUT_EXCEL_PATH a inizio file.")
    extractor = TableViewsExtractor(INPUT_EXCEL_PATH, OUTPUT_EXCEL_PATH or "")
    out_path = extractor.run()
    print(f"Viste scritte in: {out_path}")
