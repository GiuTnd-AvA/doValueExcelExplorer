# -----------------------------------------------------------------------------
# Scopo: legge il percorso di file excel da una lista presente in un file excel
# e verifica per ogni file se contiene una connessione di qualunque tipo.
# Produce un nuovo file excel con i risultati.
# -----------------------------------------------------------------------------

import os
import argparse
from typing import List

# Dependences from project
from BusinessLogic.Excel_Metadata_Extractor import ExcelMetadataExtractor
from Connection.Get_Xml_Connection import GetXmlConnection
from Connection.Get_Excel_Connection import GetXlsConnection
from Connection.Connessione_Senza_Txt import ConnessioniSenzaTxt

# Third-party for reading/writing Excel
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    import pandas as pd
except Exception:
    pd = None

# Configurazione: percorso Excel di input di default.
# Imposta qui il file da usare quando non passi --input da CLI.
# Esempio: INPUT_EXCEL_PATH = r"C:\\Percorso\\Connessioni assenti da verificare.xlsx"
INPUT_EXCEL_PATH: str | None = None


def read_paths_from_excel(input_excel: str) -> List[str]:
    """Reads file paths from Sheet1, column A of the given Excel file.

    Returns a list of non-empty strings, whitespace-stripped.
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")

    wb = load_workbook(input_excel, read_only=True, data_only=True)
    ws = wb.worksheets[0]  # Sheet1 (prima scheda)
    paths: List[str] = []
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0].value
        if cell is None:
            continue
        path = str(cell).strip()
        if path:
            paths.append(path)
    wb.close()
    return paths


def has_any_connection(file_path: str) -> bool:
    """Determine if the given Excel file contains any connection of any type.

    Strategy:
    - .xls: try GetXlsConnection (OLE streams heuristics)
    - .xlsx/.xlsm: try connections.xml via GetXmlConnection; fallback to COM via ConnessioniSenzaTxt;
      finally check metadata for xl/connections.xml presence.
    """
    if not os.path.exists(file_path):
        return False

    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".xls":
            conn = GetXlsConnection(file_path)
            conn.get_connection()
            return any([
                getattr(conn, 'server', None),
                getattr(conn, 'database', None),
                getattr(conn, 'schema', None),
                getattr(conn, 'table', None),
                getattr(conn, 'source', None)
            ])

        if ext in (".xlsx", ".xlsm"):
            # 1) connections.xml parsing
            xml = GetXmlConnection(file_path)
            infos = xml.extract_connection_info()
            if infos:
                return True

            # 2) COM workbook connections (excludes Power Query type 7)
            try:
                com = ConnessioniSenzaTxt(file_path)
                found = com.estrai_connessioni()
                if found:
                    return True
            except Exception:
                # COM not available or Excel not installed; ignore
                pass

            # 3) Fallback: metadata check for presence of xl/connections.xml
            meta = ExcelMetadataExtractor(file_path)
            meta.get_metadata(file_path)
            return meta.collegamento_esterno == 'Si'

        # Non-Excel file types: no connection
        return False
    except Exception:
        # Conservative fallback: if detection fails, assume no connection
        return False


def write_results_to_excel(output_path: str, rows: List[List[str]]):
    """Write results to a new Excel file with columns:
    Percorso | Nome File | HaConnessione (Si/No)
    """
    if pd is None:
        raise RuntimeError("pandas non installato. Installa 'pip install pandas openpyxl'.")
    df = pd.DataFrame(rows, columns=["Percorso", "NomeFile", "HaConnessione"])
    # Ensure directory exists
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name='Risultati')


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Legge i percorsi in colonna A di Sheet1 dall'Excel di input "
            "('Connessioni assenti da verificare') e produce un nuovo Excel "
            "con Percorso, Nome file e informazione (Si/No) se contiene una connessione."
        )
    )
    parser.add_argument(
        "--input",
        default=INPUT_EXCEL_PATH,
        help=(
            "Percorso dell'Excel di input (Sheet1 col A). Se omesso, "
            "usa la variabile INPUT_EXCEL_PATH definita a inizio script."
        )
    )
    parser.add_argument(
        "--output",
        default=os.path.join(os.getcwd(), "Connessioni_verificate.xlsx"),
        help="Percorso dell'Excel di output. Default: ./Connessioni_verificate.xlsx"
    )
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output

    if not input_path:
        parser.error(
            "Devi specificare --input oppure valorizzare INPUT_EXCEL_PATH all'inizio script."
        )

    paths = read_paths_from_excel(input_path)
    rows: List[List[str]] = []
    total = len(paths)
    for idx, p in enumerate(paths, start=1):
        # Normalize and expand environment variables
        fp = os.path.expandvars(os.path.expanduser(p))
        nome = os.path.basename(fp)
        has_conn = has_any_connection(fp)
        rows.append([fp, nome, 'Si' if has_conn else 'No'])
        print(f"[{idx}/{total}] {fp} -> {'Si' if has_conn else 'No'}")

    write_results_to_excel(output_path, rows)
    print(f"Report scritto: {output_path}")


if __name__ == "__main__":
    main()
