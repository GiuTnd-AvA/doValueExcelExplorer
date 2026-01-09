# -----------------------------------------------------------------------------
# Scopo: legge un file Excel con colonne:
# server, db, schema, table, Object type, DDL
# Per i record con Object type = 'view' fa l'append della DDL in un file
# aggiungendo un commento iniziale nel formato:
# -- <n> server\database\schema\table.sql
# e subito sotto la definizione della vista (DDL).
# Produce output .sql e/o .txt.
# -----------------------------------------------------------------------------

import os
from typing import Optional, List, Dict
import re

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# Config di default (puoi sovrascrivere da CLI se necessario)
# Imposta qui il percorso del tuo file Excel se vuoi evitare di passarlo da CLI.
# Esempio: DEFAULT_EXCEL_PATH = r"C:\percorso\al\file.xlsx"
DEFAULT_EXCEL_PATH: Optional[str] = None
DEFAULT_SHEET_NAME: Optional[str] = None  # usa il primo foglio se None
DEFAULT_OUTPUT_PATH: Optional[str] = None  # se None, crea output vicino all'Excel


class ViewsDDLAppender:
    """Legge un Excel e appende le DDL delle sole viste in un file unico.

    Richiede intestazioni (case-insensitive, spazi tollerati):
    - server
    - db
    - schema
    - table
    - Object type
    - DDL
    """

    REQUIRED_HEADERS = ["server", "db", "schema", "table", "object type", "ddl"]

    def __init__(self, excel_path: str, output_txt: Optional[str] = None, sheet_name: Optional[str] = DEFAULT_SHEET_NAME, create_sql_copy: bool = True):
        if load_workbook is None:
            raise RuntimeError("openpyxl non installato. Installa 'pip install openpyxl'.")
        if not excel_path:
            raise ValueError("Percorso Excel non valorizzato.")
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel non trovato: {excel_path}")
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        if output_txt:
            self.output_txt = output_txt
        else:
            base_dir = os.path.dirname(excel_path) or os.getcwd()
            self.output_txt = os.path.join(base_dir, "Views_Append.txt")
        self.create_sql_copy = create_sql_copy

    @staticmethod
    def _norm_header(h: Optional[str]) -> str:
        return (str(h).strip().lower() if h is not None else "")

    def _read_rows(self) -> List[Dict[str, str]]:
        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
        try:
            ws = wb[self.sheet_name] if self.sheet_name and self.sheet_name in wb.sheetnames else wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [self._norm_header(h) for h in (rows[0] or [])]
            # Mappa header -> indice
            idx_map: Dict[str, int] = {}
            for needed in self.REQUIRED_HEADERS:
                if needed in headers:
                    idx_map[needed] = headers.index(needed)
                else:
                    # Prova varianti senza spazi
                    needed_compact = needed.replace(" ", "")
                    found_idx = None
                    for i, h in enumerate(headers):
                        if h.replace(" ", "") == needed_compact:
                            found_idx = i
                            break
                    if found_idx is not None:
                        idx_map[needed] = found_idx
                    else:
                        raise RuntimeError(f"Colonna richiesta non trovata: '{needed}'")

            data_rows = rows[1:]  # salta intestazione
            out: List[Dict[str, str]] = []
            def extract_view_from_ddl(ddl: str) -> Dict[str, Optional[str]]:
                if not ddl:
                    return {"schema": None, "name": None}
                text = str(ddl)
                # normalize whitespace
                # Try bracketed schema and name: CREATE [OR ALTER] VIEW [schema].[name]
                m = re.search(r"\bcreate\s+(?:or\s+alter\s+)?view\s+\[([^\]]+)\]\s*\.\s*\[([^\]]+)\]",
                              text, flags=re.IGNORECASE | re.DOTALL)
                if m:
                    return {"schema": m.group(1), "name": m.group(2)}
                # Try bracketed name only: CREATE VIEW [name]
                m = re.search(r"\bcreate\s+(?:or\s+alter\s+)?view\s+\[([^\]]+)\]",
                              text, flags=re.IGNORECASE | re.DOTALL)
                if m:
                    return {"schema": None, "name": m.group(1)}
                # Try unbracketed schema.name: CREATE VIEW schema.name
                m = re.search(r"\bcreate\s+(?:or\s+alter\s+)?view\s+([a-zA-Z0-9_]+)\s*\.\s*([a-zA-Z0-9_]+)\b",
                              text, flags=re.IGNORECASE)
                if m:
                    return {"schema": m.group(1), "name": m.group(2)}
                # Try unbracketed name only
                m = re.search(r"\bcreate\s+(?:or\s+alter\s+)?view\s+([a-zA-Z0-9_]+)\b",
                              text, flags=re.IGNORECASE)
                if m:
                    return {"schema": None, "name": m.group(1)}
                return {"schema": None, "name": None}

            for r in data_rows:
                if not r:
                    continue
                def get(col: str) -> str:
                    i = idx_map[col]
                    return str(r[i]).strip() if i < len(r) and r[i] is not None else ""
                obj_type = get("object type").lower()
                if obj_type != "view":
                    continue
                server = get("server")
                db = get("db")
                schema = get("schema")
                table = get("table")
                ddl = get("ddl")
                # Fallback: se 'table' (nome vista) manca, prova a estrarlo dalla DDL
                if (not table) and ddl:
                    parsed = extract_view_from_ddl(ddl)
                    if parsed["name"]:
                        table = parsed["name"]
                    if (not schema) and parsed["schema"]:
                        schema = parsed["schema"]
                # Ultimo fallback: se ancora manca il nome, crea un placeholder stabile
                if not table:
                    # usa un nome generico per evitare intestazioni tronche
                    table = "UNKNOWN_VIEW"
                # se manca qualcosa, comunque produce riga (ddl può essere vuota)
                out.append({
                    "server": server,
                    "db": db,
                    "schema": schema,
                    "table": table,
                    "ddl": ddl,
                })
            return out
        finally:
            wb.close()

    def run(self) -> str:
        rows = self._read_rows()
        if not rows:
            raise RuntimeError("Nessuna vista trovata nell'Excel.")

        out_dir = os.path.dirname(self.output_txt)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        with open(self.output_txt, "w", encoding="utf-8", errors="ignore") as f:
            for idx, r in enumerate(rows, start=1):
                header_path = os.path.join(r["server"], r["db"], r["schema"], f"{r['table']}.sql")
                # Converti in percorso con backslash per il commento richiesto
                header_path_win = header_path.replace("/", "\\")
                f.write(f"-- {idx} {header_path_win}\n")
                if r["ddl"]:
                    f.write(str(r["ddl"]))
                f.write("\n")  # newline tra blocchi

        if self.create_sql_copy:
            sql_copy = os.path.splitext(self.output_txt)[0] + ".sql"
            try:
                with open(self.output_txt, "rb") as src, open(sql_copy, "wb") as dst:
                    dst.write(src.read())
            except Exception:
                pass

        return self.output_txt


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Append DDL delle viste da Excel in output .txt/.sql")
    # Rende l'argomento posizionale 'excel' opzionale: se non fornito, usa DEFAULT_EXCEL_PATH
    ap.add_argument("excel", nargs="?", help="Percorso al file Excel di input")
    ap.add_argument("-o", "--output", help="Percorso file di output .txt")
    ap.add_argument("-s", "--sheet", help="Nome foglio Excel da usare (default: primo)")
    ap.add_argument("--no-sql-copy", action="store_true", help="Non creare la copia .sql")
    args = ap.parse_args()

    create_sql_copy = not args.no_sql_copy
    # Calcola il percorso Excel: CLI oppure DEFAULT_EXCEL_PATH
    excel_path = args.excel or DEFAULT_EXCEL_PATH
    if not excel_path:
        raise SystemExit(
            "Errore: specifica il percorso Excel (argomento 'excel') oppure imposta DEFAULT_EXCEL_PATH all'inizio del file."
        )

    app = ViewsDDLAppender(excel_path, args.output or DEFAULT_OUTPUT_PATH, args.sheet or DEFAULT_SHEET_NAME, create_sql_copy=create_sql_copy)
    out = app.run()
    print(f"Output creato: {out}")
    if create_sql_copy:
        print(f"Copia .sql: {os.path.splitext(out)[0] + '.sql'}")
